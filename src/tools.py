from agents import RunContextWrapper
from .context import AppContext
from typing import Any, Dict, Optional, List
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from typing_extensions import TypedDict

# NOTE: ExcelManager methods return None on success; tool wrappers must not treat None as failure.
#       Always return True (or success dict) when no exception is raised.

# Define specific types for style components
class FontStyle(TypedDict, total=False):
    name: Optional[str]
    size: Optional[float]
    bold: Optional[bool]
    italic: Optional[bool]
    vertAlign: Optional[str]
    underline: Optional[str]
    strike: Optional[bool]
    color: Optional[str]

class FillStyle(TypedDict, total=False):
    fill_type: Optional[str] # e.g., 'solid'
    start_color: Optional[str]
    end_color: Optional[str]

class BorderStyleDetails(TypedDict, total=False):
    style: Optional[str] # e.g., 'thin', 'medium'
    color: Optional[str]

class BorderStyle(TypedDict, total=False):
    left: Optional[BorderStyleDetails]
    right: Optional[BorderStyleDetails]
    top: Optional[BorderStyleDetails]
    bottom: Optional[BorderStyleDetails]
    diagonal: Optional[BorderStyleDetails]
    diagonal_direction: Optional[int]
    outline: Optional[bool]
    vertical: Optional[BorderStyleDetails]
    horizontal: Optional[BorderStyleDetails]


# Define the main style structure using TypedDict
class CellStyle(TypedDict, total=False):
    """Defines the structure for cell styling options."""
    font: Optional[FontStyle]
    fill: Optional[FillStyle]
    border: Optional[BorderStyle]

# All tool functions are ready for @function_tool decoration.

# Tool: Open workbook
def open_workbook_tool(ctx: RunContextWrapper[AppContext], file_path: str) -> Any:
    """
    Opens or attaches to an Excel workbook at the given path.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        file_path (str): Path to the Excel workbook to open or attach.

    Returns:
        bool: True if the workbook was opened successfully.
        dict: {'error': str} if an error occurred.
    """
    try:
        # Delegate to ExcelManager and propagate its return value (usually None).
        return ctx.context.excel_manager.open_workbook(file_path)
    except Exception as e:
        print(f"[TOOL ERROR] open_workbook_tool: {e}")
        return {"error": f"Failed to open workbook '{file_path}': {e}"}

# Tool: Get all sheet names
def get_sheet_names_tool(ctx: RunContextWrapper[AppContext]) -> Any:
    """
    Retrieves all worksheet names in the current Excel workbook.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.

    Returns:
        List[str]: A list of worksheet names on success.
        dict: {'error': str} if an error occurred.
    """
    try:
        return ctx.context.excel_manager.get_sheet_names()
    except Exception as e:
        # Use print for server-side logging, return dict for agent
        print(f"[TOOL ERROR] get_sheet_names_tool: {e}")
        return {"error": f"Failed to get sheet names: {e}"}

# Tool: Get active sheet name
def get_active_sheet_name_tool(ctx: RunContextWrapper[AppContext]) -> Any:
    """
    Retrieves the name of the currently active worksheet.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.

    Returns:
        str: Name of the active sheet on success.
        dict: {'error': str} if an error occurred.
    """
    try:
        return ctx.context.excel_manager.get_active_sheet_name()
    except Exception as e:
        print(f"[TOOL ERROR] get_active_sheet_name_tool: {e}")
        return {"error": f"Failed to get active sheet name: {e}"}

# Tool: Set cell value
def set_cell_value_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, cell_address: str, value: Any) -> Any:
    """
    Sets the value of a single cell. **Use this tool only once per turn; if two + cells need updates, call `set_cell_values_tool` instead.**

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        sheet_name (str): Name of the worksheet.
        cell_address (str): Cell address in A1 notation (e.g., 'B2').
        value (Any): The value to set in the cell (number, text, date, or formula).

    Returns:
        bool: True if the cell was updated successfully.
        dict: {'error': str} if an error occurred.
    """
    print(f"[TOOL] set_cell_value_tool: {sheet_name}!{cell_address} value={value}")
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'set_cell_value_tool' failed: 'sheet_name' cannot be empty."}
    if not cell_address:
        return {"error": "Tool 'set_cell_value_tool' failed: 'cell_address' cannot be empty."}
    # Note: Validating 'value: Any' is complex; rely on underlying function for now.
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.set_cell_value(sheet_name, cell_address, value)
    except Exception as e:
        print(f"[TOOL ERROR] set_cell_value_tool: {e}")
        return {"error": f"Exception setting cell value for {sheet_name}!{cell_address}: {e}"}

# Tool: Get cell value
def get_cell_value_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, cell_address: str) -> Any:
    print(f"[TOOL] get_cell_value_tool: {sheet_name}!{cell_address}")
    """
    Retrieves the value from a single cell.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        sheet_name (str): Name of the worksheet.
        cell_address (str): Cell address in A1 notation (e.g., 'C3').

    Returns:
        Any: The cell value (None if empty).
        dict: {'error': str} if an error occurred.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'get_cell_value_tool' failed: 'sheet_name' cannot be empty."}
    if not cell_address:
        return {"error": "Tool 'get_cell_value_tool' failed: 'cell_address' cannot be empty."}
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.get_cell_value(sheet_name, cell_address)
    except Exception as e:
        print(f"[TOOL ERROR] get_cell_value_tool: {e}")
        return {"error": f"Exception getting cell value for {sheet_name}!{cell_address}: {e}"}

def get_range_values_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, range_address: str) -> Any:
    """
    Retrieves values from a rectangular cell range using the unified ExcelManager.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        sheet_name (str): Name of the worksheet.
        range_address (str): Excel range in A1 notation (e.g., 'A1:C5').

    Returns:
        List[List[Any]]: 2-D array of values on success.
        dict: {'error': str} if an error occurred.
    """
    print(f"[TOOL] get_range_values_tool: {sheet_name}!{range_address}")
    # Input validation
    if not sheet_name:
        return {"error": "Tool 'get_range_values_tool' failed: 'sheet_name' cannot be empty."}
    if not range_address:
        return {"error": "Tool 'get_range_values_tool' failed: 'range_address' cannot be empty."}
    try:
        return ctx.context.excel_manager.get_range_values(sheet_name, range_address)
    except Exception as e:
        print(f"[TOOL ERROR] get_range_values_tool: {e}")
        return {"error": f"Exception getting range values for {sheet_name}!{range_address}: {e}"}
# Tool: Set range style
# The SDK automatically handles JSON conversion to the Pydantic/TypedDict model.
def set_range_style_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, range_address: str, style: CellStyle) -> Any:
    print(f"[TOOL] set_range_style_tool: {sheet_name}!{range_address} style={style}")
    """
    Applies formatting styles to a cell range.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        sheet_name (str): Name of the worksheet.
        range_address (str): Excel range in A1 notation (e.g., 'A1:B2').
        style (CellStyle): Dictionary defining 'font', 'fill', and 'border' styles.

    Returns:
        bool: True if styles applied successfully.
        dict: {'error': str} if an error occurred.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'set_range_style_tool' failed: 'sheet_name' cannot be empty."}
    if not range_address:
        return {"error": "Tool 'set_range_style_tool' failed: 'range_address' cannot be empty."}
    if not style: # Check if the style dictionary itself is empty
        return {"error": "Tool 'set_range_style_tool' failed: 'style' dictionary cannot be empty."}
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.set_range_style(sheet_name, range_address, style)
    except Exception as e:
        print(f"[TOOL ERROR] set_range_style_tool: {e}")
        return {"error": f"Exception applying cell style to {sheet_name}!{range_address}: {e}"}

# Tool: Create sheet
def create_sheet_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, index: Optional[int] = None) -> Any:
    print(f"[TOOL] create_sheet_tool: sheet_name={sheet_name}, index={index}")
    """
    Creates a new sheet with the given name and optional index.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the new sheet.
        index: Optional position for the new sheet.
    Returns:
        True if successful, or an error message.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'create_sheet_tool' failed: 'sheet_name' cannot be empty."}
    # Optional: Add check for invalid characters in sheet names if needed, though openpyxl might handle this.
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.create_sheet(sheet_name, index)
    except Exception as e:
        print(f"[TOOL ERROR] create_sheet_tool: {e}")
        return {"error": f"Exception creating sheet '{sheet_name}': {e}"}

# Tool: Delete sheet
def delete_sheet_tool(ctx: RunContextWrapper[AppContext], sheet_name: str) -> Any:
    print(f"[TOOL] delete_sheet_tool: sheet_name={sheet_name}")
    """
    Deletes the specified sheet from the workbook.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the sheet to delete.
    Returns:
        True if successful, or an error message.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'delete_sheet_tool' failed: 'sheet_name' cannot be empty."}
    # --- End Validation ---
    try:
        # Delegate deletion and propagate its return (usually None).
        return ctx.context.excel_manager.delete_sheet(sheet_name)
    except Exception as e:
        print(f"[TOOL ERROR] delete_sheet_tool: {e}")
        return {"error": f"Exception deleting sheet '{sheet_name}': {e}"}

# Tool: Merge cells
def merge_cells_range_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, range_address: str) -> Any:
    print(f"[TOOL] merge_cells_range_tool: {sheet_name}!{range_address}")
    """
    Merges a range of cells in the specified sheet.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the sheet.
        range_address: Range to merge (e.g., 'A1:B2').
    Returns:
        True if successful, or an error message.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'merge_cells_range_tool' failed: 'sheet_name' cannot be empty."}
    if not range_address:
        return {"error": "Tool 'merge_cells_range_tool' failed: 'range_address' cannot be empty."}
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.merge_cells_range(sheet_name, range_address)
    except Exception as e:
        print(f"[TOOL ERROR] merge_cells_range_tool: {e}")
        return {"error": f"Exception merging cells {sheet_name}!{range_address}: {e}"}

# Tool: Unmerge cells
def unmerge_cells_range_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, range_address: str) -> Any:
    print(f"[TOOL] unmerge_cells_range_tool: {sheet_name}!{range_address}")
    """
    Unmerges a range of cells in the specified sheet.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the sheet.
        range_address: Range to unmerge (e.g., 'A1:B2').
    Returns:
        True if successful, or an error message.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'unmerge_cells_range_tool' failed: 'sheet_name' cannot be empty."}
    if not range_address:
        return {"error": "Tool 'unmerge_cells_range_tool' failed: 'range_address' cannot be empty."}
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.unmerge_cells_range(sheet_name, range_address)
    except Exception as e:
        print(f"[TOOL ERROR] unmerge_cells_range_tool: {e}")
        return {"error": f"Exception unmerging cells {sheet_name}!{range_address}: {e}"}

# Tool: Set row height
def set_row_height_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, row_number: int, height: float) -> Any:
    print(f"[TOOL] set_row_height_tool: {sheet_name} row {row_number} height={height}")
    """
    Sets the height of a row in the specified sheet.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the sheet.
        row_number: Row number (1-based).
        height: Height in points.
    Returns:
        True if successful, or an error message.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'set_row_height_tool' failed: 'sheet_name' cannot be empty."}
    if not isinstance(row_number, int) or row_number <= 0:
        return {"error": f"Tool 'set_row_height_tool' failed: 'row_number' must be a positive integer (got {row_number})."}
    if not isinstance(height, (int, float)) or height < 0:
         # Excel might allow 0 height, but negative is invalid.
        return {"error": f"Tool 'set_row_height_tool' failed: 'height' must be a non-negative number (got {height})."}
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.set_row_height(sheet_name, row_number, height)
    except Exception as e:
        print(f"[TOOL ERROR] set_row_height_tool: {e}")
        return {"error": f"Exception setting row height for row {row_number} in '{sheet_name}': {e}"}

# Tool: Set column width
def set_column_width_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, column_letter: str, width: float) -> Any:
    print(f"[TOOL] set_column_width_tool: {sheet_name} column {column_letter} width={width}")
    """
    Sets the width of a column in the specified sheet.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the sheet.
        column_letter: Column letter (e.g., 'A').
        width: Width in points.
    Returns:
        True if successful, or an error message.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'set_column_width_tool' failed: 'sheet_name' cannot be empty."}
    if not column_letter or not isinstance(column_letter, str):
        return {"error": "Tool 'set_column_width_tool' failed: 'column_letter' must be a non-empty string."}
    if not isinstance(width, (int, float)) or width < 0:
        # Excel might allow 0 width, but negative is invalid.
        return {"error": f"Tool 'set_column_width_tool' failed: 'width' must be a non-negative number (got {width})."}
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.set_column_width(sheet_name, column_letter.upper(), width)
    except Exception as e:
        print(f"[TOOL ERROR] set_column_width_tool: {e}")
        return {"error": f"Exception setting column width for column {column_letter.upper()} in '{sheet_name}': {e}"}

# Tool: Set cell formula
def set_cell_formula_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, cell_address: str, formula: str) -> Any:
    print(f"[TOOL] set_cell_formula_tool: {sheet_name}!{cell_address} formula={formula}")
    """
    Sets a formula in the specified cell. Formula must start with '='.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the sheet.
        cell_address: Cell address (e.g., 'A1').
        formula: The formula string (with or without '=' prefix).
    Returns:
        True if successful, or an error message.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"error": "Tool 'set_cell_formula_tool' failed: 'sheet_name' cannot be empty."}
    if not cell_address:
        return {"error": "Tool 'set_cell_formula_tool' failed: 'cell_address' cannot be empty."}
    if not formula: # Formula string cannot be empty
        return {"error": "Tool 'set_cell_formula_tool' failed: 'formula' cannot be empty."}
    # --- End Validation ---
    try:
        return ctx.context.excel_manager.set_cell_formula(sheet_name, cell_address, formula)
    except Exception as e:
        print(f"[TOOL ERROR] set_cell_formula_tool: {e}")
        return {"error": f"Exception setting cell formula for {sheet_name}!{cell_address}: {e}"}

# Tool: Set multiple cell values
class CellValueMap(TypedDict):
    """A mapping from cell addresses (e.g., 'A1') to values of any type (number, text, date, etc.)."""
    # This TypedDict represents a dictionary like: {"A1": "value1", "B2": "value2"}
    # It doesn't need explicit fields defined here because it acts as a type hint
    # for arbitrary key-value pairs where keys are strings (cell addresses)
    # and values are strings (cell values). The actual validation happens
    # during runtime or type checking based on how it's used.
    pass # Use pass instead of ellipsis for an empty body


class SetCellValuesResult(TypedDict, total=False):
    success: bool
    error: str


def set_cell_values_tool(
    ctx: RunContextWrapper[AppContext],
    sheet_name: str,
    data: Dict[str, Any]
) -> SetCellValuesResult:
    # --- Input Validation ---
    if not sheet_name:
        return {"success": False, "error": "Tool 'set_cell_values_tool' failed: 'sheet_name' cannot be empty."}
    if not data: # Check if data dictionary is empty
        print("[TOOL ERROR] set_cell_values_tool called with empty data dictionary.")
        return {"success": False, "error": "Tool 'set_cell_values_tool' failed: The 'data' dictionary cannot be empty. Provide cell addresses and values."}
    # Optional: Add validation for keys (cell addresses) and values if needed, though manager might handle it.
    # --- End Validation ---
    print(f"[TOOL] set_cell_values_tool: {sheet_name}, {len(data)} cells")
    """
    Sets the values of multiple cells in the specified sheet from a dictionary.
    Args:
        ctx: Agent context (injected automatically).
        sheet_name: Name of the sheet.
        data: Dictionary mapping cell addresses (e.g., 'A1') to values of any supported Excel type.
    Returns:
        A result dict with 'success' True if successful, or 'error' message if failed.
    """
    try:
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return {"success": True}
    except Exception as e:
        print(f"[TOOL ERROR] set_cell_values_tool: {e}")
        return {"success": False, "error": f"Exception setting multiple cell values in '{sheet_name}': {e}"}

# ------------------------------------------------------------------ #
#  Bulk helper tools                                                 #
# ------------------------------------------------------------------ #

def set_table_tool(ctx: RunContextWrapper[AppContext],
                   sheet_name: str,
                   top_left: str,
                   rows: List[List[Any]]) -> Any:
    """
    Bulk-write a 2-D python list into *sheet_name* starting at *top_left* (e.g. 'A2').
    Saves ≥30 single calls on header+data tables.
    """
    if not sheet_name or not top_left or not rows:
        return {"error": "sheet_name, top_left, rows are required"}
    try:
        col_letter, r0 = coordinate_from_string(top_left)
        c0_idx = column_index_from_string(col_letter)
        data = {
            f"{get_column_letter(c0_idx + c)}{r0 + r}": v
            for r, row in enumerate(rows)
            for c, v in enumerate(row)
        }
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return True
    except Exception as e:
        print(f"[TOOL ERROR] set_table_tool: {e}")
        return {"error": f"Bulk write table in '{sheet_name}' starting at '{top_left}' failed: {e}"}

# Tool: Insert a formatted Excel table
def insert_table_tool(
    ctx: RunContextWrapper[AppContext],
    sheet_name: str,
    start_cell: str,
    columns: List[Any],
    rows: List[List[Any]],
    table_name: Optional[str] = None,
    table_style: Optional[str] = None,
) -> Any:
    """
    Inserts a formatted Excel table (ListObject) into the worksheet.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        sheet_name (str): Worksheet to insert the table into.
        start_cell (str): Top-left cell for the table in A1 notation (e.g., 'A1').
        columns (List[Any]): Header names for each column.
        rows (List[List[Any]]): Data rows matching the headers.
        table_name (Optional[str]): Optional name for the Excel table.
        table_style (Optional[str]): Optional Excel table style (e.g., 'TableStyleMedium2').

    Returns:
        bool: True if table inserted successfully.
        dict: {'error': str} if an error occurred.
    """
    try:
        ctx.context.excel_manager.insert_table(sheet_name, start_cell, columns, rows, table_name, table_style)
        return True
    except Exception as e:
        print(f"[TOOL ERROR] insert_table_tool: {e}")
        return {"error": f"Exception in insert_table_tool: {e}"}

# Tool: Bulk-write rows starting at column A
def set_rows_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, start_row: int, rows: List[List[Any]]) -> Any:
    """
    Writes a 2-D Python list *rows* into *sheet_name* beginning at **column A**
    and **start_row**.  Each inner list is written to consecutive columns.

    Args:
        sheet_name (str): Worksheet name.
        start_row   (int): First row number (1-based).
        rows  (List[List[Any]]): List of rows to write.

    Returns:
        True on success, or {'error': str} on failure.
    """
    if not sheet_name:
        return {"error": "Tool 'set_rows_tool' failed: 'sheet_name' cannot be empty."}
    if not isinstance(start_row, int) or start_row <= 0:
        return {"error": "Tool 'set_rows_tool' failed: 'start_row' must be a positive integer."}
    if not rows:
        return {"error": "Tool 'set_rows_tool' failed: 'rows' cannot be empty."}

    try:
        data = {}
        for r_idx, row_vals in enumerate(rows):
            row_no = start_row + r_idx
            for c_idx, val in enumerate(row_vals):
                addr = f"{get_column_letter(c_idx + 1)}{row_no}"
                data[addr] = val
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return True
    except Exception as e:
        print(f"[TOOL ERROR] set_rows_tool: {e}")
        return {"error": f"Bulk write rows in '{sheet_name}' starting at row {start_row} failed: {e}"}

# Tool: Bulk-write columns starting at row 1
def set_columns_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, start_col: str, cols: List[List[Any]]) -> Any:
    """
    Writes a 2-D Python list *cols* into *sheet_name* beginning at **row 1**
    and **start_col** (column letter). Each inner list becomes a column written downward.

    Args:
        sheet_name (str): Worksheet name.
        start_col  (str): Column letter where the first column should be written.
        cols (List[List[Any]]): List of columns; each column is a list of values.

    Returns:
        True on success, or {'error': str} on failure.
    """
    if not sheet_name:
        return {"error": "Tool 'set_columns_tool' failed: 'sheet_name' cannot be empty."}
    if not start_col or not isinstance(start_col, str):
        return {"error": "Tool 'set_columns_tool' failed: 'start_col' must be a column letter."}
    if not cols:
        return {"error": "Tool 'set_columns_tool' failed: 'cols' cannot be empty."}

    try:
        c0_idx = column_index_from_string(start_col.upper())
        data = {}
        for c_idx, col_vals in enumerate(cols):
            col_letter = get_column_letter(c0_idx + c_idx)
            for r_idx, val in enumerate(col_vals):
                addr = f"{col_letter}{r_idx + 1}"
                data[addr] = val
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return True
    except Exception as e:
        print(f"[TOOL ERROR] set_columns_tool: {e}")
        return {"error": f"Bulk write columns in '{sheet_name}' starting at column {start_col} failed: {e}"}

# Tool: Bulk-write disjoint named ranges
def set_named_ranges_tool(ctx: RunContextWrapper[AppContext],
                          sheet_name: str,
                          mapping: Dict[str, Any]) -> Any:
    """
    Writes to multiple named ranges in one call.

    Args:
        sheet_name (str): Any existing sheet in the workbook (kept for symmetry).
        mapping    (Dict[str, Any]): {range_name: scalar | list | 2-D list}

    Returns:
        True on success or {'error': str}.
    """
    if not mapping:
        return {"error": "Tool 'set_named_ranges_tool' failed: 'mapping' cannot be empty."}

    try:
        book = ctx.context.excel_manager.book  # xlwings Book object
        for rng_name, val in mapping.items():
            try:
                rng = book.names[rng_name].refers_to_range
            except KeyError:
                return {"error": f"Named range '{rng_name}' not found."}
            rng.value = val
        return True
    except Exception as e:
        print(f"[TOOL ERROR] set_named_ranges_tool: {e}")
        return {"error": f"Failed to set named ranges: {e}"}

# Tool: Copy-paste range (values | formulas | formats)
def copy_paste_range_tool(
    ctx: RunContextWrapper[AppContext],
    src_sheet: str,
    src_range: str,
    dst_sheet: str,
    dst_anchor: str,
    paste_opts: str,
) -> Any:
    """
    Clone *src_range* from *src_sheet* and paste-special into *dst_sheet*
    beginning at *dst_anchor*.

    Args:
        src_sheet   (str): Source worksheet name.
        src_range   (str): Source range (A1 style).
        dst_sheet   (str): Destination worksheet name.
        dst_anchor  (str): Top-left destination cell (A1 style).
        paste_opts  (str): 'values' | 'formulas' | 'formats'.

    Returns:
        True on success or {'error': str}.
    """
    if not all([src_sheet, src_range, dst_sheet, dst_anchor]):
        return {"error": "All sheet/range parameters are required."}
    if paste_opts.lower() not in {"values", "formulas", "formats"}:
        return {"error": "paste_opts must be 'values', 'formulas', or 'formats'."}
    try:
        ctx.context.excel_manager.copy_paste_range(
            src_sheet, src_range, dst_sheet, dst_anchor, paste_opts
        )
        return True
    except Exception as e:
        print(f"[TOOL ERROR] copy_paste_range_tool: {e}")
        return {"error": f"Failed to copy/paste range: {e}"}

def set_columns_widths_tool(ctx: RunContextWrapper[AppContext],
                            sheet_name: str,
                            widths: Dict[str, float]) -> Any:
    """Set multiple column widths in one call (openpyxl column_dimensions)."""
    try:
        for col, w in widths.items():
            ctx.context.excel_manager.set_column_width(sheet_name, col, w)
        return True
    except Exception as e:
        print(f"[TOOL ERROR] set_columns_widths_tool: {e}")
        return {"error": f"Exception setting column widths in '{sheet_name}': {e}"}

def set_range_formula_tool(ctx: RunContextWrapper[AppContext],
                           sheet_name: str,
                           range_address: str,
                           template: str) -> Any:
    """
    Apply *template* row-wise to the leftmost cell in each row of *range_address*.
    Example: template='=SUM(B{row}:E{row})' on F2:F6.
    """
    # Input validation
    if not sheet_name:
        return {"error": "Tool 'set_range_formula_tool' failed: 'sheet_name' cannot be empty."}
    if not range_address:
        return {"error": "Tool 'set_range_formula_tool' failed: 'range_address' cannot be empty."}
    try:
        start_cell, end_cell = range_address.split(":")
        col_letter, row_start = coordinate_from_string(start_cell)
        _, row_end = coordinate_from_string(end_cell)
        for r in range(row_start, row_end + 1):
            address = f"{col_letter}{r}"
            ctx.context.excel_manager.set_cell_formula(sheet_name, address, template.format(row=r))
        return True
    except Exception as e:
        print(f"[TOOL ERROR] set_range_formula_tool: {e}")
        return {"error": f"Exception applying range formula for {sheet_name}!{range_address}: {e}"}

# ------------------------------------------------------------------ #
#  Composite write-and-verify tool                                   #
# ------------------------------------------------------------------ #

class WriteVerifyResult(TypedDict, total=False):
    success: bool
    diff: Dict[str, Any]

def write_and_verify_range_tool(
    ctx: RunContextWrapper[AppContext],
    sheet_name: str,
    data: Dict[str, str],
) -> WriteVerifyResult:
    """
    Writes multiple cells and verifies the write by reading back the values.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        sheet_name (str): Name of the worksheet.
        data (Dict[str, Any]): Mapping of cell addresses to expected values.

    Returns:
        dict: {'success': True} if all values match.
              {'success': False, 'diff': dict} where 'diff' maps each mismatched cell
              to {'expected': Any, 'actual': Any}.
    """
    if not sheet_name:
        return {"success": False, "diff": {"error": "'sheet_name' cannot be empty."}}
    if not data:
        return {"success": False, "diff": {"error": "'data' dictionary cannot be empty."}}

    # 1. Write
    try:
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
    except Exception as e:
        return {"success": False, "diff": {"error": f"Write failed: {e}"}}

    # 2. Verify
    diff: Dict[str, Any] = {}
    for addr, expected in data.items():
        try:
            actual = ctx.context.excel_manager.get_cell_value(sheet_name, addr)
            if actual != expected:
                diff[addr] = {"expected": expected, "actual": actual}
        except Exception as e:
            diff[addr] = {"expected": expected, "actual": f"(read-error: {e})"}

    if diff:
        return {"success": False, "diff": diff}
    return {"success": True}

# Tool: Get full sheet as a structured "dataframe"
def get_dataframe_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, header: bool = True) -> Any:
    """
    Returns the entire sheet as a structured dump:
        {"columns": [...], "rows": [[...], ...]}
    Args:
        sheet_name: Worksheet name.
        header: Treat first row as headers (default True).
    """
    print(f"[TOOL] get_dataframe_tool: sheet={sheet_name}, header={header}")
    if not sheet_name:
        return {"error": "Tool 'get_dataframe_tool' failed: 'sheet_name' cannot be empty."}
    try:
        return ctx.context.excel_manager.get_sheet_dataframe(sheet_name, header)
    except Exception as e:
        print(f"[TOOL ERROR] get_dataframe_tool: {e}")
        return {"error": f"Exception dumping sheet '{sheet_name}': {e}"}

# ------------------------------------------------------------------ #
#  Style-inspection tools                                            #
# ------------------------------------------------------------------ #

def get_cell_style_tool(
    ctx: RunContextWrapper[AppContext], sheet_name: str, cell_address: str
) -> Any:
    """
    Return the style dict (font/fill/border) for a single cell.
    """
    try:
        return ctx.context.excel_manager.get_cell_style(sheet_name, cell_address)
    except Exception as e:
        print(f"[TOOL ERROR] get_cell_style_tool: {e}")
        return {"error": f"Failed to get cell style for {sheet_name}!{cell_address}: {e}"}

def get_range_style_tool(
    ctx: RunContextWrapper[AppContext], sheet_name: str, range_address: str
) -> Any:
    """
    Return a mapping of cell_address -> style_dict for a rectangular range.
    """
    try:
        return ctx.context.excel_manager.get_range_style(sheet_name, range_address)
    except Exception as e:
        print(f"[TOOL ERROR] get_range_style_tool: {e}")
        return {"error": f"Failed to get range style for {sheet_name}!{range_address}: {e}"}

# ------------------------------------------------------------------ #
#  New: Snapshot and revert tools                                    #
# ------------------------------------------------------------------ #
def snapshot_tool(ctx: RunContextWrapper[AppContext]) -> Any:
    """
    Saves a temporary snapshot of the current workbook state.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.

    Returns:
        dict: {'snapshot_path': str} path to the saved snapshot file.
        dict: {'error': str} if an error occurred.
    """
    try:
        path = ctx.context.excel_manager.snapshot()
        return {"snapshot_path": path}
    except Exception as e:
        print(f"[TOOL ERROR] snapshot_tool: {e}")
        return {"error": f"Failed to take snapshot: {e}"}

def revert_snapshot_tool(ctx: RunContextWrapper[AppContext]) -> Any:
    """
    Reverts the workbook to the last snapshot taken.
    """
    try:
        # Propagate manager return value (None on success).
        return ctx.context.excel_manager.revert_to_snapshot()
    except Exception as e:
        print(f"[TOOL ERROR] revert_snapshot_tool: {e}")
        return {"error": f"Failed to revert to snapshot: {e}"}

# ------------------------------------------------------------------ #
#  (Existing) Save workbook tool                                     #
# ------------------------------------------------------------------ #
# Tool: Save workbook
def save_workbook_tool(ctx: RunContextWrapper[AppContext], file_path: str) -> Any:
    """
    Saves the current workbook to the given file path.

    Args:
        ctx (RunContextWrapper[AppContext]): Agent context containing the ExcelManager.
        file_path (str): Destination file path where the workbook should be saved.

    Returns:
        bool: True if saved successfully.
        dict: {'error': str} if an error occurred.
    """
    print(f"[TOOL] save_workbook_tool: path={file_path}")
    # --- Input Validation ---
    if not file_path:
        return {"error": "Tool 'save_workbook_tool' failed: 'file_path' cannot be empty."}
    # --- End Validation ---
    try:
        ctx.context.excel_manager.save_workbook(file_path)
        return True
    except Exception as e:
        print(f"[TOOL ERROR] save_workbook_tool: {e}")
        return {"error": f"Exception saving workbook to '{file_path}': {e}"}