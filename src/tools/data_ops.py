# src/tools/data_ops.py
from agents import RunContextWrapper, function_tool
from ..context import AppContext
from ..excel_ops import ExcelConnectionError
from .core_defs import ToolResult, CellValue, CellValueMap, SetCellValuesResult, WriteVerifyResult
from typing import Any, Optional, List, Dict, TYPE_CHECKING
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string # Moved here
import asyncio # Import asyncio

if TYPE_CHECKING:
    from ..excel_ops import ExcelManager # Avoid circular import

@function_tool
def set_cell_value_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, cell_address: str, value: CellValue) -> ToolResult:
    """Sets the value of a single cell in a specified worksheet.

    Use this tool for updating only *one* cell at a time. For updating multiple
    cells efficiently in a single operation, use `set_cell_values_tool`.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        cell_address: The cell address in A1 notation (e.g., 'B2', 'C5').
        value: The value to write into the cell (can be text, number, boolean, or None).

    Returns:
        ToolResult: {'success': True} if the value was set successfully.
                    {'success': False, 'error': str} on failure (e.g., sheet not found, connection error).
    """
    print(f"[TOOL] set_cell_value_tool: {sheet_name}!{cell_address} value={value}")
    # --- Input Validation ---
    if not sheet_name:
        return {"success": False, "error": "Tool 'set_cell_value_tool' failed: 'sheet_name' cannot be empty."}
    if not cell_address:
        return {"success": False, "error": "Tool 'set_cell_value_tool' failed: 'cell_address' cannot be empty."}
    # Note: Validating 'value: Any' is complex; rely on underlying function for now.
    # --- End Validation ---
    try:
        ctx.context.excel_manager.set_cell_value(sheet_name, cell_address, value)
        return {"success": True} # Explicit success
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] set_cell_value_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] set_cell_value_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] set_cell_value_tool: {e}")
        return {"success": False, "error": f"Exception setting cell value for {sheet_name}!{cell_address}: {e}"}

@function_tool
def get_cell_value_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, cell_address: str) -> ToolResult:
    """Retrieves the value from a single specified cell.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        cell_address: The cell address in A1 notation (e.g., 'C3').

    Returns:
        ToolResult: {'success': True, 'data': CellValue} where 'data' is the value read
                    from the cell (can be text, number, boolean, or None if the cell is empty).
                    {'success': False, 'error': str} on failure (e.g., sheet not found, connection error).
    """
    print(f"[TOOL] get_cell_value_tool: {sheet_name}!{cell_address}")
    # --- Input Validation ---
    if not sheet_name:
        return {"success": False, "error": "Tool 'get_cell_value_tool' failed: 'sheet_name' cannot be empty."}
    if not cell_address:
        return {"success": False, "error": "Tool 'get_cell_value_tool' failed: 'cell_address' cannot be empty."}
    # --- End Validation ---
    try:
        value = ctx.context.excel_manager.get_cell_value(sheet_name, cell_address)
        return {"success": True, "data": value}
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] get_cell_value_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] get_cell_value_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] get_cell_value_tool: {e}")
        return {"success": False, "error": f"Exception getting cell value for {sheet_name}!{cell_address}: {e}"}

@function_tool
def get_range_values_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, range_address: str) -> ToolResult:
    """Retrieves values from a rectangular range of cells.

    Reads the values from the specified range (e.g., 'A1:C5') on the given sheet.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        range_address: The cell range in A1 notation (e.g., 'A1:C5', 'B2:D10').

    Returns:
        ToolResult: {'success': True, 'data': List[List[CellValue]]} where 'data' is a 2D list
                    (list of rows, each row is a list of cell values) representing the values read.
                    Returns empty lists for empty ranges.
                    {'success': False, 'error': str} on failure (e.g., sheet not found, invalid range, connection error).
    """
    print(f"[TOOL] get_range_values_tool: {sheet_name}!{range_address}")
    # Input validation
    if not sheet_name:
        return {"success": False, "error": "Tool 'get_range_values_tool' failed: 'sheet_name' cannot be empty."}
    if not range_address:
        return {"success": False, "error": "Tool 'get_range_values_tool' failed: 'range_address' cannot be empty."}
    try:
        values = ctx.context.excel_manager.get_range_values(sheet_name, range_address)
        return {"success": True, "data": values}
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] get_range_values_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] get_range_values_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] get_range_values_tool: {e}")
        return {"success": False, "error": f"Exception getting range values for {sheet_name}!{range_address}: {e}"}

@function_tool(strict_mode=False) # Allow flexible dict structure for 'data'
def set_cell_values_tool(ctx: RunContextWrapper[AppContext],
                         sheet_name: str,
                         data: CellValueMap
                                  ) -> SetCellValuesResult: # Use specific result type alias if desired
    """Sets the values of multiple potentially non-contiguous cells efficiently.

    Use this tool to update several cells at once by providing a dictionary
    mapping cell addresses (like 'A1', 'C5') to their desired values. This is
    more efficient than calling `set_cell_value_tool` repeatedly.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        data: A dictionary where keys are cell addresses (e.g., 'A1') and values
              are the corresponding values (text, number, boolean, None) to set.

    Returns:
        SetCellValuesResult: {'success': True} if all values were set successfully.
                             {'success': False, 'error': str} on failure (e.g., sheet not found, invalid input, connection error).
                             Returns a success with a hint if the 'data' dict is empty.
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"success": False, "error": "Tool 'set_cell_values_tool' failed: 'sheet_name' cannot be empty."}
    if not data: # Check if data dictionary is empty
        print("[INFO] set_cell_values_tool: Received empty data dict – nothing to write.")
        return {
            "success": True,
            "error": None,
            "data": None,
            "hint": "No-op: empty mapping; nothing was written."
        }
    # Optional: Add validation for keys (cell addresses) and values if needed, though manager might handle it.
    # --- End Validation ---
    print(f"[TOOL] set_cell_values_tool: {sheet_name}, {len(data)} cells")
    try:
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return {"success": True}
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] set_cell_values_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] set_cell_values_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] set_cell_values_tool: {e}")
        return {"success": False, "error": f"Exception setting multiple cell values in '{sheet_name}': {e}"}

@function_tool
def set_table_tool(ctx: RunContextWrapper[AppContext],
                   sheet_name: str,
                   top_left: str,
                               rows: List[List[CellValue]]) -> ToolResult: # Changed Any to CellValue
    """Writes a 2D list of data into a sheet, starting at a specific cell.

    This tool performs a bulk write operation, placing the provided 2D list (`rows`)
    onto the worksheet (`sheet_name`) starting at the `top_left` cell address.
    It simply writes the values and does *not* create a formal Excel Table object
    (ListObject). For creating a structured Excel table with headers and styles,
    use `insert_table_tool` instead.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        top_left: The cell address (e.g., 'A1', 'B3') where the top-left corner
                  of the data block should be placed.
        rows: A 2D list (list of lists) containing the data to write. The outer list
              represents rows, and each inner list contains the cell values for that row.

    Returns:
        ToolResult: {'success': True} if the data was written successfully.
                    {'success': False, 'error': str} on failure (e.g., sheet not found, invalid input, connection error).
    """
    if not sheet_name or not top_left or not rows:
        return {"success": False, "error": "Tool 'set_table_tool' failed: sheet_name, top_left, and non-empty rows are required"}
    # Check if it's a list of lists. Type checking CellValue elements is complex at runtime,
    # rely on the type hint and API validation primarily.
    if not isinstance(rows, list) or not all(isinstance(row, list) for row in rows):
         return {"success": False, "error": "Tool 'set_table_tool' failed: 'rows' must be a list of lists."}

    print(f"[TOOL] set_table_tool: Writing {len(rows)} rows starting at {sheet_name}!{top_left}")
    try:
        col_letter, r0 = coordinate_from_string(top_left)
        c0_idx = column_index_from_string(col_letter)
        data = {
            f"{get_column_letter(c0_idx + c)}{r0 + r}": v
            for r, row in enumerate(rows)
            for c, v in enumerate(row)
        }
        # Use set_cell_values for the actual writing
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return {"success": True}
    except ExcelConnectionError as ce: # Catch if set_cell_values fails connection
        print(f"[TOOL ERROR] set_table_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] set_table_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] set_table_tool: {e}")
        return {"success": False, "error": f"Bulk write table in '{sheet_name}' starting at '{top_left}' failed: {e}"}

@function_tool
def insert_table_tool(
    ctx: RunContextWrapper[AppContext],
    sheet_name: str,
    start_cell: str,
    columns: List[str], # Expect list of strings for headers
    rows: List[List[CellValue]], # Expect list of lists for data rows
    table_name: Optional[str] = None,
    table_style: Optional[str] = None, # e.g., 'TableStyleMedium2'
) -> ToolResult:
    """Creates a formal Excel Table (ListObject) with headers, data, and optional styling.

    This tool inserts the provided column headers (`columns`) and data (`rows`) into
    the specified sheet (`sheet_name`), starting at `start_cell`. Crucially, it
    formats this range as an actual Excel Table object, which enables features like
    structured referencing, filtering, sorting, and applying table styles.

    Use this instead of `set_table_tool` when you need a structured Excel table,
    not just raw data written to cells.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the worksheet where the table will be created.
        start_cell: The top-left cell address (e.g., 'A1') for the table, where
                    the first header will be placed.
        columns: A list of strings representing the header names for each column.
        rows: A 2D list (list of lists) containing the data rows for the table.
              Each inner list corresponds to a row and should match the order of `columns`.
        table_name: An optional name to assign to the created Excel Table object.
                    If None, Excel will automatically generate a name (e.g., "Table1").
        table_style: An optional string specifying the desired table style
                     (e.g., 'TableStyleMedium9', 'TableStyleLight1'). Defaults to Excel's standard.

    Returns:
        ToolResult: {'success': True} if the table was created successfully.
                    {'success': False, 'error': str} on failure (e.g., sheet not found, invalid input, connection error).
    """
    # Basic input validation
    if not sheet_name:
        return {"success": False, "error": "Tool 'insert_table_tool' failed: 'sheet_name' cannot be empty."}
    if not start_cell:
        return {"success": False, "error": "Tool 'insert_table_tool' failed: 'start_cell' cannot be empty."}
    if not columns:
        return {"success": False, "error": "Tool 'insert_table_tool' failed: 'columns' list cannot be empty."}
    # Rows can be empty, resulting in a table with only a header row.

    print(f"[TOOL] insert_table_tool: Into {sheet_name}!{start_cell} Name='{table_name}' Style='{table_style}' ({len(columns)} cols, {len(rows)} rows)")
    try:
        ctx.context.excel_manager.insert_table(sheet_name, start_cell, columns, rows, table_name, table_style)
        return {"success": True}
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] insert_table_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] insert_table_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] insert_table_tool: {e}")
        # Make error more specific if possible (e.g., catch specific COM errors if manager raises them)
        return {"success": False, "error": f"Failed to insert table '{table_name}': {e}"}

@function_tool
def set_rows_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, start_row: int, rows: List[List[CellValue]]) -> ToolResult:
    """Writes multiple rows of data starting from column A of a specified row.

    This tool takes a 2D list (`rows`) and writes it to the specified worksheet
    (`sheet_name`), starting at the given `start_row` (1-based index) and always
    beginning in column A. Each inner list within `rows` represents a single row,
    and its elements are written into consecutive columns (A, B, C, ...).

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        start_row: The 1-based row number where the first row of data should be written.
        rows: A 2D list (list of lists) containing the data. Each inner list is a row.

    Returns:
        ToolResult: {'success': True} if the rows were written successfully.
                    {'success': False, 'error': str} on failure (e.g., sheet not found, invalid input, connection error).
    """
    if not sheet_name:
        return {"success": False, "error": "Tool 'set_rows_tool' failed: 'sheet_name' cannot be empty."}
    if not isinstance(start_row, int) or start_row <= 0:
        return {"success": False, "error": "Tool 'set_rows_tool' failed: 'start_row' must be a positive integer."}
    if not rows:
        return {"success": False, "error": "Tool 'set_rows_tool' failed: 'rows' cannot be empty."}
    if not isinstance(rows, list) or not all(isinstance(row, list) for row in rows):
         return {"success": False, "error": "Tool 'set_rows_tool' failed: 'rows' must be a list of lists."}

    print(f"[TOOL] set_rows_tool: Writing {len(rows)} rows starting at {sheet_name}!A{start_row}")
    try:
        data = {}
        for r_idx, row_vals in enumerate(rows):
            row_no = start_row + r_idx
            for c_idx, val in enumerate(row_vals):
                addr = f"{get_column_letter(c_idx + 1)}{row_no}" # Start at column A (index 1)
                data[addr] = val
        # Use set_cell_values for the actual writing
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return {"success": True}
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] set_rows_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] set_rows_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] set_rows_tool: {e}")
        return {"success": False, "error": f"Bulk write rows in '{sheet_name}' starting at row {start_row} failed: {e}"}

@function_tool
def set_columns_tool(ctx: RunContextWrapper[AppContext], sheet_name: str, start_col: str, cols: List[List[CellValue]]) -> ToolResult:
    """Writes multiple columns of data starting from row 1 of a specified column.

    This tool takes a 2D list (`cols`), where each inner list represents a column,
    and writes it vertically to the specified worksheet (`sheet_name`). The writing
    starts at row 1 of the given `start_col` (column letter like 'A', 'B').
    Subsequent inner lists are written to the columns immediately to the right
    (e.g., if `start_col` is 'B', the columns will be B, C, D, ...).

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        start_col: The column letter (e.g., 'A', 'C') where the first column of data
                   should be written, starting from row 1.
        cols: A 2D list where each inner list contains the values for a single column,
              to be written vertically downwards.

    Returns:
        ToolResult: {'success': True} if the columns were written successfully.
                    {'success': False, 'error': str} on failure (e.g., sheet not found, invalid input, connection error).
    """
    if not sheet_name:
        return {"success": False, "error": "Tool 'set_columns_tool' failed: 'sheet_name' cannot be empty."}
    if not start_col or not isinstance(start_col, str):
        return {"success": False, "error": "Tool 'set_columns_tool' failed: 'start_col' must be a column letter."}
    if not cols:
        return {"success": False, "error": "Tool 'set_columns_tool' failed: 'cols' cannot be empty."}
    if not isinstance(cols, list) or not all(isinstance(col, list) for col in cols):
         return {"success": False, "error": "Tool 'set_columns_tool' failed: 'cols' must be a list of lists."}

    print(f"[TOOL] set_columns_tool: Writing {len(cols)} columns starting at {sheet_name}!{start_col}1")
    try:
        c0_idx = column_index_from_string(start_col.upper())
        data = {}
        for c_idx, col_vals in enumerate(cols):
            col_letter = get_column_letter(c0_idx + c_idx)
            for r_idx, val in enumerate(col_vals):
                addr = f"{col_letter}{r_idx + 1}" # Start at row 1
                data[addr] = val
        # Use set_cell_values for the actual writing
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
        return {"success": True}
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] set_columns_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] set_columns_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] set_columns_tool: {e}")
        return {"success": False, "error": f"Bulk write columns in '{sheet_name}' starting at column {start_col} failed: {e}"}

@function_tool(strict_mode=False) # Allow flexible dict for rows
def append_table_rows_tool(
    ctx: RunContextWrapper[AppContext],
    sheet_name: str,
    table_name: str,
    rows: List[List[CellValue]],
) -> ToolResult:
    """Appends new data rows to an existing formal Excel Table (ListObject).

    This tool adds one or more data rows, provided in the `rows` list, to the
    bottom of a specified, pre-existing Excel Table (`table_name`) located on
    `sheet_name`. It attempts to use the native Excel table append functionality,
    which often preserves table formatting and automatically extends formulas down.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the worksheet containing the target Excel Table.
        table_name: The name of the existing Excel Table (ListObject) to append to.
        rows: A 2D list (list of lists) where each inner list represents a row of
              data to be appended. The number of items in each inner list should
              match the number of columns in the target table.

    Returns:
        ToolResult: {'success': True} if the rows were appended successfully.
                    {'success': True, 'data': message} if 'rows' was empty (no-op).
                    {'success': False, 'error': str} if an error occurred (e.g., table not found, sheet not found, data mismatch, connection error).
    """
    print(f"[TOOL] append_table_rows_tool: {sheet_name}!{table_name} (+{len(rows)} rows)")
    if not sheet_name:
        return {"success": False, "error": "Tool 'append_table_rows_tool' failed: 'sheet_name' cannot be empty."}
    if not table_name:
        return {"success": False, "error": "Tool 'append_table_rows_tool' failed: 'table_name' cannot be empty."}
    if not rows:
        print("[INFO] append_table_rows_tool: No rows provided to append.")
        return {"success": True, "data": "No rows provided to append."}  # Success, but did nothing.
    if not isinstance(rows, list) or not all(isinstance(row, list) for row in rows):
        return {"success": False, "error": "Tool 'append_table_rows_tool' failed: 'rows' must be a list of lists."}

    try:
        ctx.context.excel_manager.append_table_rows(sheet_name, table_name, rows)
        return {"success": True}
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] append_table_rows_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke:  # Catch table/sheet not found specifically
        print(f"[TOOL ERROR] append_table_rows_tool: Table or Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] append_table_rows_tool: {e}")
        return {"success": False, "error": f"Exception appending rows to table '{table_name}' on sheet '{sheet_name}': {e}"}


@function_tool(strict_mode=False) # Allow flexible dict structure for 'data'
def write_and_verify_range_tool(
    ctx: RunContextWrapper[AppContext],
    sheet_name: str,
    data: Dict[str, CellValue], # More specific type hint for values
) -> WriteVerifyResult:
    """Writes values to multiple cells and immediately verifies if the written values match.

    This tool first attempts to write the values provided in the `data` dictionary
    to the specified cells using `set_cell_values_tool`. Then, it reads back the values
    from those same cells and compares them against the expected values. It's useful
    for ensuring data integrity after write operations, especially complex ones.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the target worksheet.
        data: A dictionary mapping cell addresses (e.g., 'A1', 'B5') to the values
              that are intended to be written and verified.

    Returns:
        WriteVerifyResult:
            {'success': True} if the write operation succeeded and all read-back values
             matched the expected values in the `data` dictionary.
            {'success': False, 'diff': Dict[str, Any]} if the write failed, or if any
             read-back value did not match the expected value. The 'diff' dictionary
             details the discrepancies, mapping cell addresses to either
             {'expected': ExpectedValue, 'actual': ReadValue} for mismatches, or
             {'error': ErrorMessage} if the initial write or subsequent read failed.
    """
    if not sheet_name:
        return {"success": False, "diff": {"error": "'sheet_name' cannot be empty."}}
    if not data:
        return {"success": False, "diff": {"error": "'data' dictionary cannot be empty."}}

    print(f"[TOOL] write_and_verify_range_tool: Writing and verifying {len(data)} cells in {sheet_name}")

    # 1. Write
    try:
        # Use set_cell_values which handles connection checks and potential optimizations
        ctx.context.excel_manager.set_cell_values(sheet_name, data)
    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] write_and_verify_range_tool (Write Phase): Connection Error - {ce}")
        return {"success": False, "diff": {"error": f"Write failed (Connection Error): {ce}"}}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] write_and_verify_range_tool (Write Phase): Sheet not found - {ke}")
        return {"success": False, "diff": {"error": f"Write failed (Sheet not found): {ke}"}}
    except Exception as e:
        print(f"[TOOL ERROR] write_and_verify_range_tool (Write Phase): {e}")
        return {"success": False, "diff": {"error": f"Write failed: {e}"}}

    # 2. Verify
    diff: Dict[str, Any] = {}
    for addr, expected in data.items():
        try:
            # Use get_cell_value which handles connection checks
            actual = ctx.context.excel_manager.get_cell_value(sheet_name, addr)
            # Perform comparison (handle potential type differences if necessary, e.g., float precision)
            # Simple comparison for now:
            if actual != expected:
                print(f"[VERIFY FAIL] {sheet_name}!{addr}: Expected '{expected}' ({type(expected).__name__}), got '{actual}' ({type(actual).__name__})")
                diff[addr] = {"expected": expected, "actual": actual}
        except ExcelConnectionError as ce:
             print(f"[TOOL ERROR] write_and_verify_range_tool (Verify Phase): Connection Error reading {addr} - {ce}")
             diff[addr] = {"expected": expected, "actual": f"(read-error: Connection Error: {ce})"}
             # Optionally stop verification on connection error? For now, log and continue verifying others.
        except KeyError: # Sheet not found during verify (shouldn't happen if write succeeded, but safety check)
            print(f"[TOOL ERROR] write_and_verify_range_tool (Verify Phase): Sheet '{sheet_name}' not found reading {addr}")
            diff[addr] = {"expected": expected, "actual": "(read-error: Sheet not found)"}
        except Exception as e:
            print(f"[TOOL ERROR] write_and_verify_range_tool (Verify Phase): Error reading {addr} - {e}")
            diff[addr] = {"expected": expected, "actual": f"(read-error: {e})"}

    if diff:
        print(f"[TOOL] write_and_verify_range_tool: Verification failed for {len(diff)} cells.")
        return {"success": False, "diff": diff}

    print(f"[TOOL] write_and_verify_range_tool: Verification successful for {len(data)} cells.")
    return {"success": True}


@function_tool
def find_row_by_value_tool(ctx: RunContextWrapper[AppContext],
                           sheet_name: str,
                           column_letter: str,
                           value: CellValue) -> ToolResult:
    """Finds the first row containing a specific value within a given column.

    Scans the specified `column_letter` (e.g., "A", "C") on the `sheet_name` downwards
    from row 1 within the sheet's used range, looking for the first cell that matches
    the provided `value`. The search is case-insensitive when comparing strings.

    Args:
        ctx: Agent context (injected automatically).
        sheet_name: The name of the worksheet to search within.
        column_letter: The letter designation of the column to scan (e.g., "A", "B").
        value: The value (text, number, boolean, etc.) to search for within the column.

    Returns:
        ToolResult:
            {'success': True, 'data': int} where 'data' is the 1-based row number
            of the first matching cell found.
            {'success': True, 'data': 0} if the value is not found in the specified
            column within the used range.
            {'success': False, 'error': str} if an error occurred during the search
            (e.g., sheet not found, invalid column, connection error).
    """
    # --- Input Validation ---
    if not sheet_name:
        return {"success": False, "error": "Tool 'find_row_by_value_tool' failed: 'sheet_name' cannot be empty."}
    if not column_letter or not isinstance(column_letter, str):
        return {"success": False, "error": "Tool 'find_row_by_value_tool' failed: 'column_letter' must be a non-empty string."}
    # Value can be None, so no specific check here
    # --- End Validation ---

    print(f"[TOOL] find_row_by_value_tool: Searching {sheet_name}!{column_letter} for value '{value}'")
    try:
        # Define the range for the entire column (potential performance issue on huge sheets)
        # Consider limiting the scan range if performance becomes an issue, e.g., sheet.used_range.last_cell.row
        # full_column_range = f"{column_letter}:{column_letter}" # This might be slow
        # Alternative: Get used range and calculate column range within it
        sheet = ctx.context.excel_manager._require_sheet(sheet_name) # Get sheet object first
        used_range = sheet.used_range
        last_row = used_range.last_cell.row
        scan_range = f"{column_letter}1:{column_letter}{last_row}"
        print(f"[DEBUG] Scanning range {scan_range} for find_row_by_value")

        # Get values for the calculated range
        col_values_2d = ctx.context.excel_manager.get_range_values(sheet_name, scan_range)

        # Flatten the 2D list returned by get_range_values (it's a single column)
        col_vals = [row[0] if row else None for row in col_values_2d]

        search_value_str = str(value).strip().lower() if value is not None else ""

        for idx, cell_value in enumerate(col_vals):
            # Normalize cell value for comparison
            current_value_str = str(cell_value).strip().lower() if cell_value is not None else ""
            if current_value_str == search_value_str:
                found_row = idx + 1 # Convert 0-based index to 1-based row number
                print(f"[TOOL] find_row_by_value_tool: Found value '{value}' at row {found_row}")
                return {"success": True, "data": found_row}

        print(f"[TOOL] find_row_by_value_tool: Value '{value}' not found in column {column_letter}.")
        return {"success": True, "data": 0} # Return 0 if not found

    except ExcelConnectionError as ce:
        print(f"[TOOL ERROR] find_row_by_value_tool: Connection Error - {ce}")
        return {"success": False, "error": f"Connection Error: {ce}"}
    except KeyError as ke: # Catch sheet not found from _require_sheet
        print(f"[TOOL ERROR] find_row_by_value_tool: Sheet not found - {ke}")
        return {"success": False, "error": str(ke)}
    except Exception as e:
        print(f"[TOOL ERROR] find_row_by_value_tool: {e}")
        return {"success": False, "error": f"Failed to find row in {sheet_name}!{column_letter}: {e}"}