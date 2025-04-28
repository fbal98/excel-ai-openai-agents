"""
Unified ExcelManager: realtime xlwings backend with snapshot / undo support.
The public surface (method names / signatures) is preserved so existing tools
continue to work without modification.
"""

from __future__ import annotations

import os
import shutil
import tempfile
import asyncio
import logging # Ensure logging is imported
import os # Ensure os is imported
import platform
import tempfile
from typing import Any, Dict, List, Optional, TYPE_CHECKING, Tuple # Added Tuple


# Define a custom exception for connection issues
class ExcelConnectionError(RuntimeError):
    pass


import xlwings as xw
from xlwings.constants import LineStyle, BorderWeight, PasteType
import re
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

if TYPE_CHECKING:
    from .context import WorkbookShape # For type hinting

# Import consolidated helpers from core_defs
from .tools.core_defs import _to_bgr, _bgr_int_to_argb_hex, _normalise_rows


class ExcelManager:
    """Single realtime manager that always drives a visible Excel instance."""

    # --------------------------------------------------------------------------
    #  Construction / housekeeping
    # --------------------------------------------------------------------------
    def __init__(
        self,
        file_path: Optional[str] = None,
        visible: bool = True,
        *,
        kill_others: bool = False,
        attach_existing: bool = False,
        single_workbook: bool = True,
    ) -> None:
        """
        Prepare an ExcelManager.

        Parameters
        ----------
        file_path:
            Workbook path to open. If *None*, a new blank workbook will be created
            when the context is entered.
        visible:
            Whether the Excel window should be visible.
        kill_others:
            If *True*, attempt to quit all running Excel instances *before* launching
            a fresh one.  Defaults to *False* (do not disturb other sessions).
        attach_existing:
            If *True* **and** an Excel instance is already running, re-use the
            active instance instead of launching a new one.
        single_workbook:
            If *True*, automatically close all other open workbooks in the same Excel instance,
            leaving only the one managed by this class.
        """
        # Configuration only – real work happens in ``open()``.
        self._file_path = file_path
        self._visible = visible
        self._kill_others = kill_others
        self._attach_existing = attach_existing
        self._single_workbook = single_workbook

        # Handles populated later
        self.app: Optional[xw.App] = None
        self.book: Optional[xw.Book] = None

        # Tracking for snapshot / undo helper
        self._snapshot_path: Optional[str] = None
        self._attached_mode: bool = False # Flag to track if we attached to existing Excel
        self._last_known_book_name: Optional[str] = None # Store name for reconnection attempts

    def _is_connection_alive(self) -> bool:
        """Lightweight check if app and book handles seem valid."""
        logger = logging.getLogger(__name__)
        if self.app is None or self.book is None:
            logger.debug("_is_connection_alive: False (app or book is None)")
            return False
        try:
            # Check if app process exists (pid check)
            app_pid = self.app.pid
            # Check if book object is still accessible by querying its name
            book_name = self.book.name
            # Check if the book name matches the last known name (or update if first time)
            if self._last_known_book_name is None:
                self._last_known_book_name = book_name
            elif self._last_known_book_name != book_name:
                logger.warning(f"Book name mismatch: Expected '{self._last_known_book_name}', got '{book_name}'. Connection might be pointing to wrong book.")
                # Optionally treat this as a connection failure depending on strictness
                # return False
            # Simple check: is the book still listed in the app's books collection?
            # Use fullname for comparison as names might not be unique across closed/reopened books
            current_fullname = self.book.fullname
            found_in_app = False
            for wb in self.app.books:
                if wb.fullname == current_fullname:
                    found_in_app = True
                    break
            if not found_in_app:
                logger.warning(f"Book '{current_fullname}' not found in app.books collection.")
                return False

            logger.debug(f"_is_connection_alive: True (PID: {app_pid}, Book: {book_name})")
            return True
        except Exception as e:
            # Catch xw.errors.RPCError, pywintypes.com_error (Windows), generic Exception for AppleScript errors
            logger.warning(f"_is_connection_alive: False (Exception during check: {type(e).__name__} - {e})")
            return False

    def _attempt_reconnect(self) -> bool:
        """Try to find the Excel app and workbook again."""
        logger = logging.getLogger(__name__)
        logger.info("Attempting to reconnect to Excel...")

        original_path = self._file_path
        original_name = self._last_known_book_name or (os.path.basename(original_path) if original_path else None)

        logger.debug(f"Reconnect attempt: Looking for book (Name: {original_name}, Path: {original_path})")

        if not original_path and not original_name:
            logger.error("Reconnect failed: Cannot identify target workbook (no path or name known).")
            return False

        try:
            # Check running apps
            if not xw.apps:
                logger.warning("Reconnect failed: No Excel instances found running.")
                self.app = None
                self.book = None
                return False

            found_app: Optional[xw.App] = None
            found_book: Optional[xw.Book] = None

            # Iterate through running apps and their books
            for app_instance in xw.apps:
                logger.debug(f"Checking Excel instance PID: {app_instance.pid}")
                for wb in app_instance.books:
                    match = False
                    try:
                        wb_fullname = wb.fullname
                        wb_name = wb.name
                        # Prioritize matching by full path if known
                        if original_path and wb_fullname == original_path:
                            match = True
                            logger.debug(f"Found matching book by full path: {wb_fullname}")
                        # Fallback to matching by name if path unknown or didn't match
                        elif original_name and not original_path and wb_name == original_name:
                            match = True
                            logger.debug(f"Found matching book by name: {wb_name}")
                        # Handle case where path is known but might differ slightly (e.g. temp file?) - match name as last resort
                        elif original_path and original_name and wb_name == original_name:
                            match = True
                            logger.warning(f"Found book by name '{wb_name}' but path differs (Expected: '{original_path}', Found: '{wb_fullname}'). Reconnecting anyway.")

                    except Exception as e:
                        logger.debug(f"Error checking book details in PID {app_instance.pid}: {e}")
                        continue # Skip this book if querying fails

                    if match:
                        found_app = app_instance
                        found_book = wb
                        logger.info(f"Reconnect successful: Found book '{found_book.name}' in PID {found_app.pid}.")
                        break # Stop searching once found
                if found_app:
                    break # Stop searching apps

            if found_app and found_book:
                self.app = found_app
                self.book = found_book
                self._file_path = found_book.fullname # Update path
                self._last_known_book_name = found_book.name # Update name
                self._attached_mode = True # Assume we are now attached
                # Ensure book is activated
                try:
                    self.book.activate()
                    if self._visible:
                        self.app.activate(steal_focus=True)
                    logger.debug("Reconnected book activated.")
                except Exception as activate_err:
                    logger.warning(f"Failed to activate reconnected book: {activate_err}")
                return True
            else:
                logger.error(f"Reconnect failed: Could not find workbook '{original_name or original_path}' in any running Excel instance.")
                self.app = None
                self.book = None
                return False

        except Exception as e:
            logger.error(f"Unexpected error during reconnect attempt: {e}", exc_info=True)
            self.app = None
            self.book = None
            return False

    def _validate_connection(self) -> Tuple[xw.App, xw.Book]:
        """Ensure connection is live, attempting reconnect if needed. Returns (app, book) or raises ExcelConnectionError."""
        logger = logging.getLogger(__name__)
        if not self._is_connection_alive():
            if not self._attempt_reconnect():
                raise ExcelConnectionError("Connection to Excel lost and could not be re-established.")
        # We should have valid app and book handles here, but check for None just in case
        if self.app is None or self.book is None:
            # This case should theoretically be covered by the reconnect logic, but acts as a safeguard
            logger.error("_validate_connection: app or book is None even after checks/reconnect.")
            raise ExcelConnectionError("Internal error: Excel app or book handle is invalid after validation.")
        return self.app, self.book

    # Property to get current path safely
    @property
    def file_path(self) -> Optional[str]:
        """Return the current full path of the managed workbook, or None."""
        # Try to get the most up-to-date path from the book object first
        if self.book:
            try:
                return self.book.fullname
            except Exception:
                # Fallback to stored path if book query fails
                return self._file_path
        # If no book, return the path we were initialized with or None
        return self._file_path

    # Manual open/close methods (replaced __aenter__/__aexit__)
    async def open(self) -> None:
        """Initialize Excel app and workbook connection."""
        logger = logging.getLogger(__name__)
        logger.debug("Executing ExcelManager.open()...")
        if self.app is not None or self.book is not None:
            logger.warning("ExcelManager.open() called while already open. Attempting to close first.")
            await self.close() # Ensure clean state

        self.app = None
        self.book = None
        self._attached_mode = False # Reset flag on entry

        # Optionally close other Excel processes
        if self._kill_others:
            logger.info("Attempting to quit other running Excel instances...")
            killed_count = 0
            for _app in xw.apps:
                try:
                    _app.quit()
                    killed_count += 1
                except Exception as e:
                    logger.warning("Could not quit an Excel instance: %s", e)
            logger.info("Quit %d other Excel instance(s).", killed_count)

        # Optionally attach to an existing process
        if self._attach_existing and xw.apps:
            try:
                self.app = xw.apps.active
                self._attached_mode = True # Set flag
                logger.info("Attached to existing Excel instance (PID: %s)", self.app.pid)
            except Exception as e:
                logger.warning("Failed to attach to existing Excel instance: %s. Starting new instance.", e)
                self.app = None
                self._attached_mode = False # Ensure flag is False if attach fails

        # Start a new instance if needed
        if self.app is None:
            logger.info("Starting new Excel instance...")
            # Ensure add_book=False is critical here to avoid premature workbook creation
            self.app = xw.App(visible=self._visible, add_book=False)
            self._attached_mode = False # Explicitly false if we created a new app
            logger.info("New Excel instance started (PID: %s)", self.app.pid)

        # --- Workbook Handling ---
        target_file_path = self._file_path
        target_file_name = os.path.basename(target_file_path) if target_file_path else None
        logger.debug("Target file path: %s, Attached mode: %s", target_file_path, self._attached_mode)

        if self._attached_mode:
            # We are attached to an existing app. Try to find the target book or use active/add new.
            found_book = None
            if target_file_name:
                logger.debug("Attached mode: Looking for workbook '%s' in %d open books.", target_file_name, len(self.app.books))
                for wb in self.app.books:
                    if wb.name == target_file_name:
                        found_book = wb
                        logger.info("Found target workbook '%s' already open in attached instance.", target_file_name)
                        break
                if not found_book and target_file_path and os.path.exists(target_file_path):
                    # Book not found, try opening it if path exists
                    try:
                        logger.info("Opening specified workbook '%s' in attached instance...", target_file_path)
                        found_book = self.app.books.open(target_file_path)
                    except Exception as e:
                        logger.error("Failed to open specified workbook '%s' in attached instance: %s. Creating a new blank workbook instead.", target_file_path, e)
                        found_book = self.app.books.add()
                        logger.info("Added new blank workbook to attached instance as fallback.")
                elif not found_book:
                    # File path doesn't exist or wasn't specified, create new book
                    logger.info("Target workbook '%s' not found or path invalid. Creating new blank workbook.", target_file_name or "[New Workbook]")
                    found_book = self.app.books.add()
            else:
                # No file specified. Use active if available, otherwise add new.
                if self.app.books:
                    found_book = self.app.books.active # Use the currently active book
                    logger.info("Using active workbook '%s' in attached instance (no file specified).", found_book.name)
                else:
                    # No books open in the attached instance, add one.
                    logger.info("No workbooks open in attached instance. Adding a new blank workbook.")
                    found_book = self.app.books.add()

            self.book = found_book

        else:
            # We created a new app instance. Open file or add new book.
            if target_file_path and os.path.exists(target_file_path):
                logger.info("Opening specified workbook '%s' in new instance...", target_file_path)
                try:
                    self.book = self.app.books.open(target_file_path)
                except Exception as e:
                    logger.error("Failed to open specified workbook '%s' in new instance: %s. Creating a new blank workbook instead.", target_file_path, e)
                    self.book = self.app.books.add() # Fallback
                    logger.info("Added new blank workbook to new instance as fallback.")
            elif target_file_path:
                # Path specified but doesn't exist
                logger.info("Specified workbook '%s' not found. Creating a new workbook at this path (or default location if save fails).", target_file_path)
                self.book = self.app.books.add()
                # Attempt to save immediately to the intended path if specified
                try:
                    self.book.save(target_file_path)
                    logger.info("New workbook saved to '%s'", target_file_path)
                    # Update internal path reference if it was None
                    if self._file_path is None: self._file_path = target_file_path
                except Exception as save_err:
                    logger.warning("Could not save new workbook to '%s' immediately: %s. It remains unsaved.", target_file_path, save_err)
                    # Keep self._file_path as it was (might be None)
            else:
                # No path specified, create a new blank book
                # Check if default Book1 exists from app creation
                if len(self.app.books) > 0:
                    try:
                        self.book = self.app.books.active
                    except Exception:
                        self.book = self.app.books[0]
                    logger.info("Re‑using default workbook %s", getattr(self.book, "name", "unknown"))
                else:
                    logger.info("Adding new blank workbook to new instance…")
                    self.book = self.app.books.add()
                    self._file_path = None # Ensure no path is associated yet

        # Ensure the managed book is activated and visible
        if self.book:
            try:
                logger.debug("Activating workbook: %s", self.book.name)
                self.book.activate()
                logger.info("Managed workbook set to: %s (Path: %s)", self.book.name, self.book.fullname)
                # Update internal path if needed (e.g., if we opened an existing file or saved a new one)
                self._file_path = self.book.fullname
                if self._visible and hasattr(self.app, 'activate'):
                    logger.debug("Activating Excel application window.")
                    self.app.activate(steal_focus=True)
            except Exception as e:
                logger.warning("Could not activate workbook '%s': %s", self.book.name, e)

            # If single_workbook is True, close all other workbooks
            if self._single_workbook:
                current_book_fullname = None
                try:
                    current_book_fullname = self.book.fullname
                except:
                    current_book_fullname = None

                # On macOS, if workbook has no path, save it to a temp location
                if platform.system() == "Darwin" and not current_book_fullname:
                    temp_path = os.path.join(tempfile.gettempdir(), "NewWorkbook.xlsx")
                    try:
                        self.book.save(temp_path)
                        current_book_fullname = self.book.fullname
                    except:
                        logger.warning("Could not save new workbook to assign a path on macOS.")

                # Close all other books
                try:
                    for wb in list(self.app.books):
                        try:
                            if wb.fullname != current_book_fullname:
                                wb.close()
                        except Exception as e:
                            logger.warning(f"Could not close extra workbook: {e}")

                    # Re-check that our main book is still in self.app.books
                    found = False
                    for wb in self.app.books:
                        if wb.fullname == current_book_fullname:
                            self.book = wb
                            found = True
                            break

                    if not found:
                        logger.error("After cleanup, could not locate our main workbook by path. Possibly lost reference.")
                        self.book = None
                    else:
                        # re-activate
                        try:
                            self.book.activate()
                        except Exception as reacquire_err:
                            logger.error(f"Activation of reacquired workbook failed: {reacquire_err}")

                except Exception as close_others_err:
                    logger.error(f"Error during single_workbook cleanup: {close_others_err}")
                    # Depending on severity, might want to re-raise or just log

        else:
            # This case should ideally not happen if the logic above is correct
            raise RuntimeError("Failed to obtain a workbook handle within ExcelManager.open().")

        # Ensure at least one sheet exists in the managed book
        if self.book and not self.book.sheets:
            logger.info("Workbook '%s' has no sheets. Adding default sheet 'Sheet1'.", self.book.name)
            self.book.sheets.add(name="Sheet1") # Give it a default name

        # Store the name after successful open/attach
        if self.book:
            self._last_known_book_name = self.book.name
            self._file_path = self.book.fullname # Ensure path is also updated

        logger.debug("ExcelManager.open() completed.")

    async def close(self) -> None:
        """Gracefully close the managed workbook and potentially the Excel app."""
        logger = logging.getLogger(__name__)
        logger.debug("Executing ExcelManager.close() (Attached mode was: %s)", self._attached_mode)
        was_attached = self._attached_mode # Capture state before cleanup
        try:
            # --- Close the managed Workbook ---
            if self.book:
                book_name = "Unknown"
                book_fullname = None
                try:
                    # Get details before attempting close
                    book_name = self.book.name
                    book_fullname = self.book.fullname

                    # Check if the book is still valid/open within the app
                    # Use fullname for a more reliable check within the app's books collection
                    book_still_open = False
                    if self.app:
                        for wb in self.app.books:
                            try:
                                if wb.fullname == book_fullname:
                                    book_still_open = True
                                    break
                            except Exception as e:
                                # Handle cases where querying wb.fullname might fail for some reason
                                logger.debug(f"Could not check fullname for a workbook: {e}")

                    if book_still_open:
                        logger.info("Closing managed workbook: %s (%s)", book_name, book_fullname)
                        # Close without saving changes unless explicitly handled elsewhere (e.g., by save tool)
                        self.book.close() # Removed save_changes argument
                        logger.info("Workbook '%s' closed.", book_name)
                    else:
                        logger.warning("Managed workbook '%s' seems to be already closed or app is unavailable.", book_name)
                except Exception as e:
                    logger.error("Error closing workbook '%s': %s", book_name, e)
                finally:
                    self.book = None # Clear handle regardless
                    self._file_path = None # Clear associated path
            else:
                logger.debug("close(): No workbook handle to close.")
        finally:
            # --- Quit/Kill the Excel Application ---
            if self.app and not was_attached:
                # Only quit/kill the app if *we* created it AND there are no other books open (unless kill_others was true initially)
                should_quit_app = True
                if not self._kill_others and len(self.app.books) > 0:
                    logger.info("Not quitting Excel instance (PID: %s) as other workbooks are open and kill_others was not set.", self.app.pid)
                    should_quit_app = False

                if should_quit_app:
                    app_pid = self.app.pid
                    logger.info("Quitting Excel instance (PID: %s) as it was created by this manager and is now empty or kill_others was set.", app_pid)
                    try:
                        # Use kill when available to prevent zombie COM hosts
                        if hasattr(self.app, "kill"):
                            logger.debug("Using app.kill() for PID: %s", app_pid)
                            self.app.kill()
                        else:
                            logger.debug("Using app.quit() for PID: %s", app_pid)
                            self.app.quit()
                        logger.info("Excel instance (PID: %s) quit/killed.", app_pid)
                    except Exception as e:
                        logger.error("Error quitting/killing Excel app (PID: %s): %s", app_pid, e)
                    finally:
                        self.app = None # Clear handle regardless
                else:
                    # We are leaving the app running, just clear our handle
                    self.app = None

            elif self.app and was_attached:
                logger.info("Leaving attached Excel instance (PID: %s) running.", self.app.pid)
                # Clear the handle, but don't quit the app
                self.app = None
            else:
                # App handle might already be None if creation failed or already cleaned up
                logger.debug("close(): No app handle to clean up or already cleaned.")

            # Reset internal state
            self._attached_mode = False
            self._snapshot_path = None # Clear snapshot path on close
            self._last_known_book_name = None # Clear known name on close
        logger.debug("ExcelManager.close() completed.")

    # Optional synchronous helper for legacy call‑sites (calls async close)
    # Note: Running async code synchronously like this can be problematic.
    # It's better to call the async close from an async context.
    # This is kept for minimal compatibility but might be removed later.
    def close_sync(self) -> None:
        """Synchronously closes the Excel manager. Use async close() where possible."""
        logger = logging.getLogger(__name__)
        logger.warning("Using close_sync() is discouraged. Please switch to async close().")
        try:
            # Get or create an event loop
            loop = asyncio.get_event_loop()
            if loop.is_running():
                # If a loop is already running (e.g., in Jupyter), create a task
                # This is still not ideal, but better than loop.run_until_complete
                logger.debug("close_sync(): Loop already running, creating task for close()")
                loop.create_task(self.close())
                # Note: We can't easily wait for completion here without blocking,
                # so this is fire-and-forget in an already running loop.
            else:
                logger.debug("close_sync(): Running close() in a new event loop.")
                loop.run_until_complete(self.close())
        except RuntimeError as e:
            # Handle cases like "Cannot run the event loop while another loop is running"
            logger.error(f"Error running close_sync(): {e}. Manual cleanup might be needed.")
            # Fallback cleanup (might not be fully effective)
            self.book = None
            self.app = None
            self._attached_mode = False
            self._snapshot_path = None
        except Exception as e:
            logger.error(f"Unexpected error in close_sync(): {e}")
            # Fallback cleanup
            self.book = None
            self.app = None
            self._attached_mode = False
            self._snapshot_path = None

    # --------------------------------------------------------------------------
    #  Snapshot / undo helpers
    # --------------------------------------------------------------------------
    def snapshot(self) -> str:
        """Save a temp copy that can be rolled back to with `revert_to_snapshot()`."""
        logger = logging.getLogger(__name__)
        app, book = self._validate_connection() # Ensure connection before proceeding
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        book.save(tmp_path) # Use validated book handle
        self._snapshot_path = tmp_path
        return tmp_path

    def revert_to_snapshot(self) -> None:
        """Close current book and reopen the last snapshot (if any)."""
        logger = logging.getLogger(__name__)
        if not self._snapshot_path or not os.path.exists(self._snapshot_path):
            raise RuntimeError("No snapshot available to revert to.")
        app, book = self._validate_connection() # Ensure connection
        # Close without saving
        book.close(save_changes=False) # Use validated book handle
        # Open the snapshot using the validated app handle
        self.book = app.books.open(self._snapshot_path)
        self._last_known_book_name = self.book.name # Update name after opening snapshot
        self._file_path = self.book.fullname # Update path

    # --------------------------------------------------------------------------
    #  Ensure changes are applied
    # --------------------------------------------------------------------------
    async def ensure_changes_applied(self) -> None:
        """Asynchronously flush Excel UI and calculation pipelines.

        This method yields to the event loop for ≈0.5 s, preventing the hard
        stop caused by ``time.sleep`` while Excel finishes painting.
        """
        logger = logging.getLogger(__name__)
        try:
            app, book = self._validate_connection() # Ensure connection first
            # Force a visual and calculation refresh
            app.screen_updating = False
            self.app.screen_updating = True
            try:
                self.app.calculate()
            except Exception as calc_err:
                logger.debug(f"self.app.calculate() failed (ignored): {calc_err}")

            # Re‑activate active sheet to nudge UI
            active_sheet = self.book.sheets.active
            active_sheet.activate()

            # Give Excel a brief moment without blocking the loop
            await asyncio.sleep(0.5)
            logger.debug("Excel display refreshed.")
        except Exception as e:
            logger.debug(f"Could not refresh Excel display: {e}")

    # Removed unused save_with_confirmation method

    # --------------------------------------------------------------------------
    #  Explicit save helpers
    # --------------------------------------------------------------------------
    def save_workbook(self, file_path: Optional[str] = None) -> str:
        """Save the current workbook. If no path is provided, save to a default location. Returns the saved path."""
        logger = logging.getLogger(__name__)
        app, book = self._validate_connection() # Ensure connection

        target_path = file_path or self._file_path # Use provided path, then internal, then generate default

        if not target_path:
            # Generate a default filename with timestamp if no path given or stored
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            # Try saving in current directory first
            target_path = f"workbook_{timestamp}.xlsx"
            logging.debug(f"No file path provided or stored, generated default: {target_path}")

        # Ensure the path has .xlsx extension
        if not target_path.lower().endswith('.xlsx'):
            target_path += '.xlsx'

        return self.save_as(target_path) # Call save_as and return its result (the path)

    def save_as(self, file_path: str) -> str:
        """Save the workbook to the specified path, ensuring proper extension. Returns the saved path."""
        logger = logging.getLogger(__name__)
        app, book = self._validate_connection() # Ensure connection

        # Ensure the path has .xlsx extension
        if not file_path.lower().endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            self.book.save(file_path)
            self._file_path = self.book.fullname # Update internal path after successful save
            logger.info(f"Workbook saved successfully to: {file_path}")
            return file_path
        except Exception as e:
            logger.error(f"Failed to save workbook to {file_path}: {e}")
            raise RuntimeError(f"Failed to save workbook to {file_path}: {e}") from e

    # --------------------------------------------------------------------------
    #  New: open workbook helper
    # --------------------------------------------------------------------------
    def open_workbook(self, file_path: str) -> None:
        """Close the current book without saving and open the workbook at file_path."""
        # Need a valid app handle to open a book
        try:
            app, _ = self._validate_connection() # Check connection, ignore book handle for now
        except ExcelConnectionError:
            # If connection is lost AND cannot be re-established, we can't open a new book reliably.
            # Alternative: could try starting a completely new app instance here?
            logger.error("Excel connection lost, cannot open new workbook reliably.")
            raise ExcelConnectionError("Excel connection lost, cannot open new workbook.")

        logger = logging.getLogger(__name__)
        try:
            if self.book: # Use self.book here as _validate_connection might have reconnected it
                try:
                    logger.debug(f"Closing current workbook '{self.book.name}' before opening new one.")
                    self.book.close(save_changes=False)
                except Exception as close_err:
                    logger.warning(f"Could not close previous workbook: {close_err}")
                    pass # Continue trying to open the new one

            logger.info(f"Opening workbook: {file_path}")
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")

            self.book = self.app.books.open(file_path)
            self._file_path = self.book.fullname # Update internal path
            logger.info(f"Successfully opened workbook: {self.book.name} ({self._file_path})")

            # Ensure there's at least one sheet and activate the book
            if not self.book.sheets:
                logger.warning(f"Workbook '{self.book.name}' has no sheets. Adding default 'Sheet1'.")
                self.book.sheets.add(name="Sheet1")
            self.book.activate()

        except Exception as e:
            logger.error(f"Failed to open workbook {file_path}: {e}")
            # Reset handles if open fails catastrophically
            self.book = None
            self._file_path = None
            raise RuntimeError(f"Failed to open workbook {file_path}: {e}") from e
        self._snapshot_path = None # Clear snapshot path after opening a new book

    # --------------------------------------------------------------------------
    #  Basic workbook / sheet info
    # --------------------------------------------------------------------------
    def get_sheet_names(self) -> List[str]:
        logger = logging.getLogger(__name__)
        try:
            app, book = self._validate_connection()
            return [s.name for s in book.sheets]
        except ExcelConnectionError:
            return [] # Return empty list if connection lost
        except Exception as e:
            logging.error(f"Failed to get sheet names: {e}")
            return []

    def get_active_sheet_name(self) -> Optional[str]:
        logger = logging.getLogger(__name__)
        try:
            app, book = self._validate_connection()
            return book.sheets.active.name
        except ExcelConnectionError:
            logging.warning("Connection lost while trying to get active sheet name.")
            return None
        except Exception as e:
            # Catch specific xw errors if needed, e.g., if no sheet is active
            logging.warning(f"Could not get active sheet name: {e}")
            # Fallback: return the first sheet name if available
            if self.book and self.book.sheets:
                return self.book.sheets[0].name
            return None

    def quick_scan_shape(self) -> Optional[WorkbookShape]:
        """
        Scans the current workbook state via xlwings and returns a WorkbookShape object.
        Returns None if connection fails.
        Raises exceptions if critical operations fail (e.g., accessing book).
        Logs warnings for non-critical issues (e.g., cannot read headers).
        """
        from .context import WorkbookShape # Avoid circular import at top level
        import logging
        logger = logging.getLogger(__name__)

        try:
            app, book = self._validate_connection()
        except ExcelConnectionError:
            logger.error("quick_scan_shape failed: Connection to Excel lost.")
            # Decide what to return: None, empty shape, or raise? Let's return None.
            return None

        shape = WorkbookShape()
        # book already validated

        # 1. Scan sheets for used range and headers
        for sheet in book.sheets:
            try:
                sheet_name = sheet.name
                # Get used range - handle potential errors if sheet is empty
                # Cross‑platform used‑range detection (works on both Windows COM and macOS)
                used = sheet.used_range      # xlwings Range (never None)
                last_cell = used.last_cell   # xlwings Range
                last_addr = last_cell.address.replace("$", "")
                # Handle completely empty sheet where used_range might be just A1 but .value is None
                shape.sheets[sheet_name] = f"A1:{last_addr}" if used.value is not None else "A1:A1"

                # Get headers (first row) - handle potential errors/empty rows
                try:
                    # Fast path: fetch first row directly through COM to avoid many Range calls
                    # Use expand to get the full row width according to used range
                    header_range = sheet.range('A1').expand('right')
                    header_values = header_range.value

                    if isinstance(header_values, list): # More than one cell in header row
                        # Track the original length for logging
                        original_length = len(header_values)
                        # Remove trailing empty columns to reduce token usage
                        while header_values and (header_values[-1] is None or str(header_values[-1]).strip() == ""):
                            header_values.pop()
                        # Ensure all headers are strings, handle None/empty
                        shape.headers[sheet_name] = [str(c) if c is not None else "" for c in header_values]

                        # Log information about trimmed columns
                        retained = len(header_values)
                        trimmed = original_length - retained
                        if trimmed > 0:
                            logger.debug(f"Sheet '{sheet_name}': Headers trimmed from {original_length} to {retained} columns (removed {trimmed} empty trailing columns)")
                        else:
                            logger.debug(f"Sheet '{sheet_name}': Headers retained ({retained} columns).")

                    elif header_values is not None: # Handle single-cell header/sheet case
                        shape.headers[sheet_name] = [str(header_values)]
                        logger.debug(f"Sheet '{sheet_name}': Single header cell found.")
                    else: # Empty first row or empty sheet
                        shape.headers[sheet_name] = []
                        logger.debug(f"Sheet '{sheet_name}': No header values found.")

                except Exception as header_err:
                    logger.warning(f"Could not read headers for sheet '{sheet_name}': {header_err}. Defaulting to empty list.")
                    shape.headers[sheet_name] = [] # Fallback to empty list

            except Exception as sheet_err:
                logger.error(f"Error processing sheet '{getattr(sheet, 'name', 'unknown')}': {sheet_err}. Skipping sheet in shape.")
                continue # Skip this sheet on error

        # 2. Scan named ranges
        shape.names = {} # Initialize names dict
        try:
            for name_obj in book.names:
                nm = name_obj.name
                try:
                    # Check if refers_to_range exists and retrieve address
                    # This attribute tries to resolve the reference to a range object
                    refers_range = name_obj.refers_to_range
                    if refers_range:
                        # Get the full address including sheet name if possible
                        addr = refers_range.address.replace("$", "")
                        shape.names[nm] = addr
                        logger.debug(f"Resolved named range '{nm}' to address '{addr}'")
                    else:
                        # If refers_to_range fails (e.g., constant, complex formula, or error), store raw string
                        refers_to_str = getattr(name_obj, 'refers_to', '#REF!') # Get raw string safely
                        # Log as DEBUG not WARNING, as the name likely exists but resolution failed.
                        logger.debug(f"Could not resolve address for named range '{nm}' (refers_to='{refers_to_str}'). Storing raw refers_to string in shape.")
                        shape.names[nm] = refers_to_str
                except Exception as name_ref_err:
                    # Handle other errors during resolution, store raw string
                    refers_to_str = getattr(name_obj, 'refers_to', '#REF!') # Get raw string safely
                    # Log as DEBUG not WARNING
                    logger.debug(f"Error resolving named range '{nm}' (refers_to='{refers_to_str}'): {name_ref_err}. Storing raw refers_to string in shape.")
                    shape.names[nm] = refers_to_str
        except Exception as names_err:
            logger.error(f"Error accessing named ranges collection: {names_err}. Skipping named ranges in shape.")
            shape.names = {} # Ensure names is an empty dict on error
            # Continue without names if there's a general error
    
        shape.version = 0 # Base version, caller (AppContext) will manage incrementing
        return shape

    def get_sheet(self, sheet_name: str) -> Optional[xw.Sheet]:
        try:
            app, book = self._validate_connection()
            return book.sheets[sheet_name]
        except ExcelConnectionError:
            logging.error(f"Connection lost while trying to get sheet '{sheet_name}'.")
            return None
        except (KeyError, ValueError, Exception) as e: # Catch generic exceptions too
            logging.warning(f"Could not get sheet '{sheet_name}': {e}")
            return None

    # Removed unused fill_ranges method

    # --------------------------------------------------------------------------
    #  Cell value helpers
    # --------------------------------------------------------------------------
    def set_cell_value(self, sheet_name: str, cell_address: str, value: Any) -> None:
        logger = logging.getLogger(__name__)
        sheet = self._require_sheet(sheet_name)
        try:
            sheet.range(cell_address).value = value
        except Exception as e:
            logging.error(f"Failed to set value for {sheet_name}!{cell_address}: {e}")
            raise # Re-raise after logging

    def get_cell_value(self, sheet_name: str, cell_address: str) -> Any:
        logger = logging.getLogger(__name__)
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        try:
            return sheet.range(cell_address).value
        except Exception as e:
            logging.error(f"Failed to get value for {sheet_name}!{cell_address}: {e}")
            raise # Re-raise after logging

    def set_cell_values(self, sheet_name: str, data: Dict[str, Any]) -> None:
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        num_cells = len(data)
        logger = logging.getLogger(__name__)

        # Optimization: Try vectorized write for rectangular ranges > 1 cell
        if num_cells > 1:
            try:
                coords = []
                min_r, min_c = float('inf'), float('inf')
                max_r, max_c = 0, 0
                for addr in data.keys():
                    col_str, row_idx = coordinate_from_string(addr)
                    col_idx = column_index_from_string(col_str)
                    coords.append({'addr': addr, 'r': row_idx, 'c': col_idx})
                    min_r, max_r = min(min_r, row_idx), max(max_r, row_idx)
                    min_c, max_c = min(min_c, col_idx), max(max_c, col_idx)

                # Check if all cells within the bounding box are present in the input data
                expected_num_cells = (max_r - min_r + 1) * (max_c - min_c + 1)
                is_rectangular_and_dense = (num_cells == expected_num_cells)

                if is_rectangular_and_dense:
                    # Build the 2D matrix in the correct order
                    rows_count = max_r - min_r + 1
                    cols_count = max_c - min_c + 1
                    matrix = [[None] * cols_count for _ in range(rows_count)]

                    # Map (row, col) to value
                    coord_map = { (item['r'], item['c']): data[item['addr']] for item in coords }

                    # Populate the matrix using the map
                    for r_offset in range(rows_count):
                        for c_offset in range(cols_count):
                            current_r = min_r + r_offset
                            current_c = min_c + c_offset
                            # Get value from map; defaults to None if somehow missing (shouldn't happen if dense)
                            matrix[r_offset][c_offset] = coord_map.get((current_r, current_c))

                    start_addr = f"{get_column_letter(min_c)}{min_r}"
                    # No need for end_addr, start_addr and matrix shape are enough for xlwings
                    # end_addr = f"{get_column_letter(max_c)}{max_r}"
                    # range_address = f"{start_addr}:{end_addr}"

                    logger.debug(f"Using vectorized write for rectangular range starting at: {sheet_name}!{start_addr}")
                    sheet.range(start_addr).options(expand='table').value = matrix # expand='table' writes the matrix
                    return  # Vectorized write successful

            except Exception as e:
                logger.warning(f"Failed to apply vectorized optimization for set_cell_values: {e}. Falling back to iterative write.")

        # Fallback: non-rectangular, sparse, single cell, or error during optimization
        logger.debug(f"Using iterative write for {num_cells} cells in {sheet_name}")
        for addr, val in data.items():
            try:
                sheet.range(addr).value = val
            except Exception as cell_err:
                logger.error(f"Failed to set value for {sheet_name}!{addr}: {cell_err}")
                # Decide whether to continue or raise the first error encountered
                # raise cell_err # Option 1: Stop on first error
                continue # Option 2: Log and continue with other cells

    # --------------------------------------------------------------------------
    #  Range helpers
    # --------------------------------------------------------------------------
    def get_range_values(self, sheet_name: str, range_address: str) -> List[List[Any]]:
        logger = logging.getLogger(__name__)
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        try:
            vals = sheet.range(range_address).value
            # xlwings returns scalar for 1×1 range; list for 1D row/col; list-of-lists for 2D
            if vals is None:
                return [[]] # Represent empty range or single empty cell consistently
            if not isinstance(vals, list):
                # Single cell value
                return [[vals]]
            if vals and not isinstance(vals[0], list):
                # 1-D list (either a single row or single column), normalise to 2-D list-of-lists
                # Need to determine if it's a row or column based on the range object
                rng_obj = sheet.range(range_address)
                if rng_obj.shape[0] == 1: # It's a single row
                    return [vals]
                elif rng_obj.shape[1] == 1: # It's a single column
                    return [[v] for v in vals]
                else: # Should not happen if shape is consistent with value type
                    logging.warning(f"Ambiguous 1D list returned for range {range_address}. Assuming row format.")
                    return [vals] # Default assumption: it's a row
            # Already a 2-D list-of-lists or empty list for empty multi-cell range
            return vals if vals else [[]] # Ensure empty list becomes list containing one empty list
        except Exception as e:
            logging.error(f"Failed to get values for range {sheet_name}!{range_address}: {e}")
            raise # Re-raise after logging

    # _normalise_rows moved to core_defs.py

    # --------------------------------------------------------------------------
    #  Styles (minimal viable impl)
    # --------------------------------------------------------------------------
    def set_range_style(
        self, sheet_name: str, range_address: str, style: Dict[str, Any]
    ) -> None:
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)
        try:
            rng = sheet.range(range_address)

            # Font → bold, color
            if "font" in style and isinstance(style["font"], dict):
                font_style = style["font"]
                # Handle bold
                bold = font_style.get("bold")
                if bold is not None:
                    try:
                        rng.font.bold = bool(bold)
                    except Exception as e:
                        logger.warning(f"Style Error ({range_address}): Failed to set font bold: {e}")

                # Handle font color
                color_argb = font_style.get("color")
                if color_argb is not None:
                    try:
                        rng.font.color = _to_bgr(color_argb)
                    except ValueError as ve: # Catch specific color format error
                        logger.warning(f"Style Error ({range_address}): Invalid font color format '{color_argb}': {ve}")
                    except Exception as e:
                        logger.warning(f"Style Error ({range_address}): Failed to set font color: {e}")

            # Fill
            if "fill" in style and isinstance(style["fill"], dict):
                fill_style = style["fill"]
                # Get fill type (currently only handles solid color)
                # fill_type = fill_style.get("fill_type", "solid")

                # Handle start color (assuming solid fill)
                color_argb = fill_style.get("start_color")
                if color_argb is not None:
                    try:
                        rng.color = _to_bgr(color_argb) # xlwings uses .color for background fill
                    except ValueError as ve: # Catch specific color format error
                        logger.warning(f"Style Error ({range_address}): Invalid fill color format '{color_argb}': {ve}")
                    except Exception as e:
                        logger.warning(f"Style Error ({range_address}): Failed to set fill color: {e}")

            # Borders
            # Note: Applying individual borders is complex and error-prone with xlwings/COM across platforms.
            # Using rng.api.BorderAround is often more reliable for simple outlines.
            if "border" in style and isinstance(style["border"], dict):
                border_style_dict = style["border"]
                try:
                    # -- Define style mappings --
                    # Map style name to BorderWeight enum/value
                    weight_map = {
                        # String names
                        "thin": BorderWeight.thin,            # 2  
                        "medium": BorderWeight.medium,        # -4138
                        "thick": BorderWeight.thick,          # 4
                        "hairline": BorderWeight.hairline,    # 1
                        "heavy": BorderWeight.heavy,          # -4138 (same as medium)
                        # Direct integers (for compatibility)
                        "1": BorderWeight.hairline,           # 1
                        "2": BorderWeight.thin,               # 2
                        "4": BorderWeight.thick,              # 4
                        # Default to thin for unknown values
                    }
                    
                    line_style_map = { # Map to LineStyle
                        # Default to continuous for most weights
                        "thin": LineStyle.continuous,         # 1
                        "medium": LineStyle.continuous,      
                        "thick": LineStyle.continuous,
                        "hairline": LineStyle.continuous,
                        "heavy": LineStyle.continuous,
                        # Additional styles if needed:
                        "dashed": LineStyle.dashed,           # -4115
                        "dotted": LineStyle.dot,              # -4118
                        "double": LineStyle.double,           # -4119
                        # Direct integers (for compatibility)
                        "1": LineStyle.continuous,  # continuous
                        "-4115": LineStyle.dashed,  # dashed
                        "-4118": LineStyle.dot,     # dotted
                        "-4119": LineStyle.double,  # double
                    }
                    
                    # -- Handle outline vs individual sides --
                    # First check for outline specification
                    outline_spec = border_style_dict.get("outline")
                    if outline_spec:
                        # Handle outline border (either True or a dict with style/color)
                        border_to_apply = outline_spec if isinstance(outline_spec, dict) else {}
                        border_style = str(border_to_apply.get("style", "thin")).lower() # Default thin
                        color_argb = border_to_apply.get("color", "FF000000") # Default black
                        
                        # Normalize border style to Excel constants
                        xl_weight = weight_map.get(border_style, BorderWeight.thin)
                        xl_linestyle = line_style_map.get(border_style, LineStyle.continuous)
                        xl_color = _to_bgr(color_argb) # Convert color
                        
                        # Apply outline border
                        rng.api.BorderAround(
                            LineStyle=xl_linestyle,
                            Weight=xl_weight,
                            Color=xl_color
                        )
                        logger.debug(f"Applied outline border ({border_style}, {color_argb}) to {range_address}")
                    
                    # Handle individual borders if specified
                    sides = {"left": 7, "right": 10, "top": 8, "bottom": 9}  # Excel border position constants
                    
                    for side_name, border_idx in sides.items():
                        if side_name in border_style_dict:
                            side_spec = border_style_dict[side_name]
                            if not isinstance(side_spec, dict):
                                continue  # Skip if not a dict with style/color
                                
                            # Extract style info for this side
                            side_style = str(side_spec.get("style", "thin")).lower()  # Default thin
                            side_color = side_spec.get("color", "FF000000")  # Default black
                            
                            # Normalize to Excel constants
                            side_weight = weight_map.get(side_style, BorderWeight.thin)
                            side_linestyle = line_style_map.get(side_style, LineStyle.continuous)
                            side_color_val = _to_bgr(side_color)
                            
                            # Apply the individual border
                            try:
                                # Get the Borders collection and apply to specific index
                                border_obj = rng.api.Borders(border_idx)
                                border_obj.LineStyle = side_linestyle
                                border_obj.Weight = side_weight
                                border_obj.Color = side_color_val
                                logger.debug(f"Applied {side_name} border ({side_style}, {side_color}) to {range_address}")
                            except Exception as side_err:
                                logger.warning(f"Failed to apply {side_name} border to {range_address}: {side_err}")

                except ValueError as ve: # Catch specific color format error
                    logger.warning(f"Style Error ({range_address}): Invalid border color format: {ve}")
                except Exception as e:
                    logger.warning(f"Style Error ({range_address}): Failed to apply border: {e}")

            # Alignment
            if "alignment" in style and isinstance(style["alignment"], dict):
                alignment = style["alignment"]
                horiz = alignment.get("horizontal")
                if horiz is not None:
                    # Use xlwings constants for alignment where possible
                    alignment_map = {
                        "general": xw.constants.HAlign.xlHAlignGeneral,
                        "left": xw.constants.HAlign.xlHAlignLeft,
                        "center": xw.constants.HAlign.xlHAlignCenter,
                        "right": xw.constants.HAlign.xlHAlignRight,
                        "fill": xw.constants.HAlign.xlHAlignFill,
                        "justify": xw.constants.HAlign.xlHAlignJustify,
                        "centercontinuous": xw.constants.HAlign.xlHAlignCenterAcrossSelection,
                        "distributed": xw.constants.HAlign.xlHAlignDistributed,
                    }
                    try:
                        rng.api.HorizontalAlignment = alignment_map.get(str(horiz).lower(), xw.constants.HAlign.xlHAlignGeneral)
                    except Exception as e:
                        logger.warning(f"Style Error ({range_address}): Failed to set horizontal alignment '{horiz}': {e}")

                vert = alignment.get("vertical")
                if vert is not None:
                    vertical_map = {
                        "top": xw.constants.VAlign.xlVAlignTop,
                        "center": xw.constants.VAlign.xlVAlignCenter,
                        "bottom": xw.constants.VAlign.xlVAlignBottom,
                        "justify": xw.constants.VAlign.xlVAlignJustify,
                        "distributed": xw.constants.VAlign.xlVAlignDistributed,
                    }
                    try:
                        rng.api.VerticalAlignment = vertical_map.get(str(vert).lower(), xw.constants.VAlign.xlVAlignCenter)
                    except Exception as e:
                        logger.warning(f"Style Error ({range_address}): Failed to set vertical alignment '{vert}': {e}")

                wrap = alignment.get("wrap_text")
                if wrap is not None:
                    try:
                        rng.api.WrapText = bool(wrap)
                    except Exception as e:
                        logger.warning(f"Style Error ({range_address}): Failed to set wrap text: {e}")

            # Number Format
            if "number_format" in style and style["number_format"] is not None:
                num_format = style["number_format"]
                try:
                    rng.number_format = str(num_format)
                except Exception as e:
                    logger.warning(f"Style Error ({range_address}): Failed to set number format '{num_format}': {e}")

            # Force update the Excel application to show changes (consider rate limiting this)
            # try:
            #     if self.app:
            #         self.app.screen_updating = False
            #         self.app.screen_updating = True
            # except Exception:
            #     pass # Ignore errors during screen update toggle
        except Exception as outer_e:
            logger.error(f"Unexpected error applying style to {sheet_name}!{range_address}: {outer_e}")

    # --------------------------------------------------------------------------
    #  Sheet management
    # --------------------------------------------------------------------------
    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> None:
        logger = logging.getLogger(__name__)
        if not self.book: raise RuntimeError("Cannot create sheet: No active workbook.")
        if sheet_name in self.get_sheet_names():
            logger.error(f"Sheet '{sheet_name}' already exists.")
            raise ValueError(f"Sheet '{sheet_name}' already exists.")
        try:
            before_sheet = None
            after_sheet = None
            num_sheets = len(self.book.sheets)

            if index is not None:
                if index < 0: # Handle negative index like Python lists
                    index = num_sheets + index
                index = max(0, min(index, num_sheets)) # Clamp index within valid range [0, num_sheets]

                if index < num_sheets:
                    before_sheet = self.book.sheets[index] # Add before this existing sheet
                # If index == num_sheets, add at the end (default behavior if before/after are None)
                # xlwings add() uses 'before' or 'after' parameter.
                # To add at index `i`, we add *before* the sheet currently at index `i`.
                # If index is 0, add before the first sheet.
                # If index is num_sheets, add after the last sheet (equivalent to before=None).

            logger.info(f"Creating sheet '{sheet_name}' at index {index if index is not None else 'end'}")
            # Pass the sheet object to 'before' parameter
            new_sheet = self.book.sheets.add(name=sheet_name, before=before_sheet, after=None)
            # new_sheet.activate() # Optionally activate the new sheet

        except Exception as e:
            logger.error(f"Failed to create sheet '{sheet_name}': {e}")
            # Don't raise generic RuntimeError, let caller handle specific exceptions if needed
            # Or wrap in a custom CreateSheetError? For now, log and let propagate.
            raise

    def delete_sheet(self, sheet_name: str) -> None:
        app, book = self._validate_connection() # Ensure connection
        logger = logging.getLogger(__name__)
        # Prevent deleting the last sheet
        if len(book.sheets) <= 1:
            logger.error(f"Cannot delete sheet '{sheet_name}': It is the only sheet remaining.")
            raise ValueError("Cannot delete the last remaining sheet.")

        sheet = self._require_sheet(sheet_name) # Raises error if not found
        alerts_enabled = True # Default to True in case app access fails
        try:
            logger.warning(f"Disabling alerts before deleting sheet '{sheet_name}'.")
            # Disable alerts to prevent confirmation dialogs
            if self.app:
                alerts_enabled = self.app.display_alerts
                self.app.display_alerts = False
            logger.info(f"Deleting sheet: {sheet_name}")
            sheet.delete()
            logger.info(f"Sheet '{sheet_name}' deleted successfully.")
        except Exception as e:
            logger.error(f"Failed to delete sheet '{sheet_name}': {e}")
            raise RuntimeError(f"Failed to delete sheet '{sheet_name}': {e}") from e
        finally:
            # Restore alerts setting
            if self.app:
                logger.warning("Re-enabling alerts.")
                self.app.display_alerts = alerts_enabled

    # --------------------------------------------------------------------------
    #  Merge / unmerge
    # --------------------------------------------------------------------------
    def merge_cells_range(self, sheet_name: str, range_address: str) -> None:
        """Merge cells in *range_address* with macOS-safe fallbacks.

        Excel on macOS occasionally throws ``OSERROR: -50 (Parameter error)``
        when the high-level ``Range.merge()`` helper is used, even for valid
        ranges.  We:

        1. Skip work if the area is already merged.
        2. Attempt the normal `rng.merge()`.
        3. If that fails, *unmerge first* (clears half-merged artefacts) and
           retry with the low-level ``rng.api.Merge()`` which accepts no
           arguments and is more tolerant on macOS.
        """
        sheet = self._require_sheet(sheet_name)  # ensures connection
        logger = logging.getLogger(__name__)
        rng = sheet.range(range_address)

        # 0️⃣ Already merged? Nothing to do.
        try:
            if (
                hasattr(rng, "merge_area")
                and rng.merge_area.address.replace("$", "") == rng.address.replace("$", "")
            ):
                logger.debug("Range %s!%s already merged – skipping.", sheet_name, range_address)
                return
        except Exception:
            # merge_area may not exist in every build; ignore
            pass

        def _attempt(label: str, func) -> bool:
            try:
                func()
                logger.debug("Merge via %s succeeded for %s!%s", label, sheet_name, range_address)
                return True
            except Exception as exc:
                logger.debug("Merge via %s failed for %s!%s: %s", label, sheet_name, range_address, exc)
                return False

        # 1️⃣ primary path
        if _attempt("rng.merge()", lambda: rng.merge()):
            return

        # 2️⃣ fallback – unmerge first, then low-level merge
        try:
            rng.unmerge()
        except Exception:
            pass  # best-effort

        if _attempt("rng.api.Merge()", lambda: rng.api.Merge()):
            return

        # Everything failed
        raise RuntimeError(f"Could not merge cells {sheet_name}!{range_address}: all attempts failed.")

    def unmerge_cells_range(self, sheet_name: str, range_address: str) -> None:
        """Unmerge cells in *range_address*.

        The previous implementation relied on ``rng.cells[0]`` which is not
        available on certain macOS builds.  We now trust Excel’s own resilience:
        calling unmerge on an already-unmerged range is a harmless no-op.
        """
        sheet = self._require_sheet(sheet_name)  # ensures connection
        logger = logging.getLogger(__name__)
        rng = sheet.range(range_address)

        # First attempt – high-level helper
        try:
            rng.unmerge()
            logger.debug("Unmerged %s!%s via rng.unmerge()", sheet_name, range_address)

            # ── Preserve value across the newly split cells ───────────────
            top_val = rng.value  # after unmerge this is now a 2-D list
            if not isinstance(top_val, list):
                top_val = [[top_val]]
            if top_val and top_val[0]:
                seed = top_val[0][0]
                if seed not in (None, ""):
                    blanks = {}
                    for cell in rng:
                        if cell.value in (None, ""):
                            blanks[cell.address.replace("$", "")] = seed
                    if blanks:
                        self.set_cell_values(sheet_name, blanks)
                        logger.debug("Replicated merged value to %d empty cell(s)", len(blanks))

            return
        except Exception as exc1:
            logger.debug("rng.unmerge() failed: %s; trying rng.api.UnMerge()", exc1)

        # Fallback – low-level API
        try:
            rng.api.UnMerge()
            logger.debug("Unmerged %s!%s via rng.api.UnMerge()", sheet_name, range_address)
        except Exception as exc2:
            logger.error("Failed to unmerge %s!%s: %s", sheet_name, range_address, exc2)
            raise RuntimeError(f"Could not unmerge cells {sheet_name}!{range_address}: {exc2}") from exc2

    # --------------------------------------------------------------------------
    #  Row / column sizing
    # --------------------------------------------------------------------------
    def set_row_height(self, sheet_name: str, row_number: int, height: Optional[float]) -> None:
        """Set the height of a specific row. If height is None, autofit."""
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)
        try:
            row_range_addr = f"{row_number}:{row_number}"
            row_range = sheet.range(row_range_addr)
            if height is not None:
                if height <= 0: # Set to hidden if height is 0 or negative
                    logger.debug(f"Setting row {row_number} on sheet '{sheet_name}' to hidden.")
                    row_range.api.EntireRow.Hidden = True
                else:
                    logger.debug(f"Setting row height for {row_number} on sheet '{sheet_name}' to {height}.")
                    row_range.row_height = height
                    row_range.api.EntireRow.Hidden = False # Ensure row is visible
            else:
                logger.debug(f"Autofitting row {row_number} on sheet '{sheet_name}'.")
                row_range.autofit() # Use autofit method on range
                # Alternative: row_range.api.EntireRow.AutoFit()
        except Exception as e:
            logger.error(f"Failed to set row height for row {row_number} in '{sheet_name}': {e}")
            # Decide whether to raise
            raise

    def set_column_width(self, sheet_name: str, column_letter: str, width: Optional[float]) -> None:
        """Set the width of a specific column. If width is None, autofit."""
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)
        col_letter_upper = column_letter.upper()
        try:
            col_range_addr = f"{col_letter_upper}:{col_letter_upper}"
            col_range = sheet.range(col_range_addr)
            # Wrap the entire block in try-catch to avoid "EntireColumn" errors
            try:
                if width is not None:
                    # If width <= 0, we skip "hidden" logic and just do nothing
                    if width <= 0:
                        logger.debug(f"Unable to hide column {col_letter_upper} due to environment limitations; ignoring.")
                    else:
                        logger.debug(f"Setting column width for {col_letter_upper} on sheet '{sheet_name}' to {width}.")
                        col_range.column_width = width
                else:
                    logger.debug(f"Autofitting column {col_letter_upper} on sheet '{sheet_name}'.")
                    col_range.autofit()
            except Exception as e:
                logger.warning(f"Could not set or autofit column width for '{sheet_name}'!{col_letter_upper}: {e}")
                # Alternative: col_range.api.EntireColumn.AutoFit()
        except Exception as e:
            logger.error(f"Failed to set column width for column {col_letter_upper} in '{sheet_name}': {e}")
            # Decide whether to raise
            raise

    # --------------------------------------------------------------------------
    #  Copy / Paste range helper
    # --------------------------------------------------------------------------
    def copy_paste_range(
        self,
        src_sheet_name: str,
        src_range: str,
        dst_sheet_name: str,
        dst_anchor: str,
        paste_opts: str = "values",
    ) -> None:
        """
        Clone *src_range* from *src_sheet_name* and paste into *dst_sheet_name*
        at *dst_anchor* (top-left cell of destination).

        paste_opts:
            • "all"     → Values, Formulas, Formats, etc. (like Ctrl+C, Ctrl+V)
            • "values"    → Values only
            • "formulas" → Formulas only
            • "formats"  → Formats only
            • "values_number_formats" -> Values and number formats
            • "column_widths" -> Column widths only (destination range size ignored)
        """
        # Validate connection before getting sheets
        app, book = self._validate_connection()
        src_sheet = self._require_sheet(src_sheet_name) # Uses validated connection implicitly
        dst_sheet = self._require_sheet(dst_sheet_name) # Uses validated connection implicitly
        logger = logging.getLogger(__name__)

        try:
            src_rng = src_sheet.range(src_range)
            dst_rng = dst_sheet.range(dst_anchor)
            
            # Safely get the destination top-left cell without assuming .cells attribute
            # First try accessing cells directly, but fall back to the range itself if needed
            try:
                # Try the normal approach first (works in most xlwings versions)
                if hasattr(dst_rng, "cells") and len(dst_rng.cells) > 0:
                    dst_rng_anchor = dst_rng.cells[0]
                    logger.debug(f"Using cells[0] for destination anchor at {dst_rng_anchor.address}")
                else:
                    # If .cells isn't available, use the range itself (for single cell)
                    dst_rng_anchor = dst_rng
                    logger.debug(f"Using range directly for destination anchor at {dst_rng_anchor.address}")
            except Exception as cell_err:
                # Fallback to using the range itself if any access issue with .cells
                logger.debug(f"Falling back to direct range for anchor due to: {cell_err}")
                dst_rng_anchor = dst_rng

            opts = paste_opts.lower()
            logger.debug(f"Copying from {src_sheet_name}!{src_range} to {dst_sheet_name}!{dst_anchor} with option '{opts}'")

            # Use API Copy/PasteSpecial for more control and reliability
            try:
                src_rng.api.Copy()
            except Exception as copy_err:
                logger.warning(f"Direct Copy() API call failed: {copy_err}. Attempting alternative approach...")
                
                # Fallback for non-COM environments: manual copy of values
                if opts == "values":
                    # Get source values and calculate target range dimensions
                    src_values = src_rng.value
                    if not isinstance(src_values, list):
                        # Handle single cell case
                        src_values = [[src_values]]
                    elif src_values and not isinstance(src_values[0], list):
                        # Handle 1D list (single row or column)
                        if src_rng.shape[0] == 1:
                            # Single row
                            src_values = [src_values]
                        else:
                            # Single column
                            src_values = [[v] for v in src_values]
                    
                    # Write values directly
                    rows = len(src_values)
                    cols = len(src_values[0]) if rows > 0 and src_values[0] else 0
                    dst_range = dst_sheet.range(dst_anchor).resize(rows, cols)
                    dst_range.value = src_values
                    logger.info(f"Fallback: Manual copy of values from {src_sheet_name}!{src_range} to {dst_sheet_name}!{dst_anchor}")
                    return  # Exit early after using fallback approach
                else:
                    # Re-raise for other paste types that need COM copy/paste
                    raise copy_err

            paste_type_map = {
                "all": PasteType.all, # -4104, xlPasteAll
                "values": PasteType.values, # -4163, xlPasteValues
                "formats": PasteType.formats, # -4122, xlPasteFormats
                "formulas": PasteType.formulas, # -4123, xlPasteFormulas
                "values_number_formats": PasteType.values_and_number_formats, # 12, xlPasteValuesAndNumberFormats
                "column_widths": PasteType.column_widths, # 8, xlPasteColumnWidths
                # Add others from xw.constants.PasteType if needed
            }

            if opts in paste_type_map:
                xl_paste_option = paste_type_map[opts]
                
                try:
                    # Try PasteSpecial API call
                    dst_rng_anchor.api.PasteSpecial(Paste=xl_paste_option)
                except Exception as paste_err:
                    logger.warning(f"PasteSpecial API call failed: {paste_err}. Attempting fallback...")
                    
                    # Fallback for specific paste types
                    if opts == "values":
                        # Try direct value assignment if API paste fails
                        src_values = src_rng.value
                        if src_values is not None:
                            # Ensure we have target range with correct dimensions
                            if isinstance(src_values, list):
                                rows = len(src_values)
                                if rows > 0 and isinstance(src_values[0], list):
                                    cols = len(src_values[0])
                                else:
                                    # Single row or not a list
                                    if src_rng.shape[0] == 1:
                                        # It's a single row
                                        rows, cols = 1, len(src_values)
                                        src_values = [src_values]
                                    else:
                                        # It's a single column
                                        rows, cols = len(src_values), 1
                                        src_values = [[v] for v in src_values]
                            else:
                                # Single cell value
                                rows, cols = 1, 1
                                src_values = [[src_values]]
                                
                            # Resize destination range and assign values
                            dst_range = dst_sheet.range(dst_anchor).resize(rows, cols)
                            dst_range.value = src_values
                            logger.info(f"Fallback: Manual assignment of values for {rows}x{cols} range")
                        else:
                            logger.warning("Source range value is None, nothing to paste")
                    else:
                        # No fallback available for other paste types
                        raise paste_err
            else:
                # Handle invalid paste option
                logger.error(f"Invalid paste_opts '{paste_opts}'. Use one of: {list(paste_type_map.keys())}")
                raise ValueError(f"Invalid paste_opts '{paste_opts}'.")

            # Clear clipboard marquee (optional, cosmetic)
            try:
                if self.app:
                    self.app.api.CutCopyMode = False
            except Exception as clear_err:
                logger.debug(f"Could not clear copy mode: {clear_err}")

        except Exception as e:
            logger.error(f"Failed to copy/paste range: {e}")
            # Ensure cut/copy mode is turned off if an error occurs mid-process
            try:
                if self.app: self.app.api.CutCopyMode = False
            except:
                pass
            raise RuntimeError(f"Failed copy/paste from {src_sheet_name}!{src_range} to {dst_sheet_name}!{dst_anchor}: {e}") from e

    # --------------------------------------------------------------------------
    #  (Currently stub) advanced APIs
    # --------------------------------------------------------------------------
    def set_cell_formula(self, sheet_name: str, cell_address: str, formula: str) -> None:
        """Sets the formula for a cell. Ensures it starts with '='."""
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)
        try:
            if not formula.startswith("="):
                formula = "=" + formula
            logger.debug(f"Setting formula for {sheet_name}!{cell_address}: {formula}")
            sheet.range(cell_address).formula = formula
        except Exception as e:
            logger.error(f"Failed to set formula for {sheet_name}!{cell_address}: {e}")
            raise # Re-raise after logging

    # --------------------------------------------------------------------------
    #  Style inspectors
    # --------------------------------------------------------------------------
    def get_cell_style(self, sheet_name: str, cell_address: str) -> Dict[str, Any]:   # noqa: D401
        """Return a minimal style dict (bold + fill color) for a single cell."""
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        try:
            return _safe_cell_style(sheet.range(cell_address))
        except Exception as e:
            logging.error(f"Failed to get style for {sheet_name}!{cell_address}: {e}")
            return {} # Return empty dict on error

    def get_range_style(self, sheet_name: str, range_address: str) -> Dict[str, Dict[str, Any]]:   # noqa: D401
        """
        Return {cell_address: style_dict} for every cell in the range (minimal style set).
            Filters out cells with default/empty style.
        """
        logger = logging.getLogger(__name__)
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        styles = {}
        try:
            rng = sheet.range(range_address)
            # Optimize for larger ranges by fetching properties in bulk if possible (platform dependent)
            # Simple iteration approach:
            for cell in rng:
                style = _safe_cell_style(cell)
                if style: # Only include if style is not empty
                    styles[cell.address.replace("$", "")] = style
            return styles
        except Exception as e:
            logging.error(f"Failed to get styles for range {sheet_name}!{range_address}: {e}")
            return {} # Return empty dict on error

    # Data-frame style dump for inspection / verification
    def get_sheet_dataframe(self, sheet_name: str, header: bool = True) -> Dict[str, list]:
        """Reads the used range of a sheet and returns data as {'columns': [], 'rows': []}."""
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)
        try:
            # Use used_range which is generally reliable
            used_rng = sheet.used_range
            # Handle case where sheet is completely empty
            if used_rng.address == '$A$1' and used_rng.value is None:
                logger.debug(f"Sheet '{sheet_name}' appears empty based on used_range.")
                return {"columns": [], "rows": []}

            # Get values; handle potential single cell case by get_range_values normalization
            full_range_addr = used_rng.address.replace('$', '')
            values = self.get_range_values(sheet_name, full_range_addr) # Ensures 2D list

            if not values or not values[0]: # Check if values is empty or first row is empty
                logger.debug(f"Sheet '{sheet_name}' used range {full_range_addr} returned no data.")
                return {"columns": [], "rows": []}

            if header:
                # Take the first row as header, ensure strings, generate names for empty headers
                columns = []
                first_row = values[0]
                for i, c in enumerate(first_row):
                    col_name = str(c) if c is not None else f"Column{i+1}"
                    # Ensure unique column names (simple append index if duplicate)
                    # original_name = col_name
                    # count = 1
                    # while col_name in columns:
                    #     col_name = f"{original_name}_{count}"
                    #     count += 1
                    columns.append(col_name) # Basic uniqueness handling might be needed for robustness

                rows = values[1:]
            else:
                # No header row, generate default column names
                num_cols = len(values[0])
                columns = [f"Column{i+1}" for i in range(num_cols)]
                rows = values

            # Normalize rows to match header width
            # Note: _normalise_rows is now imported from core_defs
            rows = _normalise_rows(columns, rows)

            logger.debug(f"Read {len(rows)} data rows with {len(columns)} columns from sheet '{sheet_name}'.")
            return {"columns": columns, "rows": rows}

        except Exception as e:
            logger.error(f"Failed to get sheet dataframe for '{sheet_name}': {e}")
            # Decide whether to raise or return empty
            return {"columns": [], "rows": []} # Return empty on error

    # --------------------------------------------------------------------------
    #  Helpers
    # --------------------------------------------------------------------------
    def _require_sheet(self, sheet_name: str) -> xw.Sheet:
        """Gets a sheet object, raising KeyError if not found or ExcelConnectionError if connection fails."""
        logger = logging.getLogger(__name__)
        app, book = self._validate_connection() # Raises ExcelConnectionError if fails
        try:
            sheet = book.sheets[sheet_name]
            return sheet
        except (KeyError, ValueError, Exception) as e: # Catch potential xlwings errors too
            available_sheets = []
            try:
                available_sheets = [s.name for s in book.sheets] # Get current sheets if possible
            except Exception:
                pass # Ignore if getting sheets also fails
            logging.error(f"Sheet '{sheet_name}' not found or inaccessible. Available: {available_sheets}. Error: {e}")
            raise KeyError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}") from e

    # --------------------------------------------------------------------------
    #  Table insertion / modification
    # --------------------------------------------------------------------------
    def insert_table(
        self,
        sheet_name: str,
        start_cell: str,
        columns: List[Any],
        rows: List[List[Any]],
        table_name: Optional[str] = None,
        table_style: Optional[str] = "TableStyleMedium9", # Default Excel style
    ) -> None:
        """
        Inserts data and formats it as an Excel table (ListObject) into the worksheet.
        """
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)
        logger.info(f"Inserting table '{table_name or 'Unnamed'}' into {sheet_name}!{start_cell} with {len(columns)} columns and {len(rows)} rows.")

        try:
            header_cell = sheet.range(start_cell)
            # Ensure rows data is normalized to correct width BEFORE calculating range
            # Note: _normalise_rows is now imported from core_defs
            normalized_rows = _normalise_rows(columns, rows)

            total_rows = 1 + len(normalized_rows) # 1 for header
            total_cols = len(columns)
            if total_cols == 0:
                logger.warning("Cannot insert table with zero columns.")
                return # Or raise error

            # Define the full range for the table data + header
            table_range = header_cell.resize(total_rows, total_cols)
            table_address = table_range.address.replace('$', '')
            logger.debug(f"Table data range: {table_address}")

            # --- Pre-check for existing data/objects ---
            # Check if the target range overlaps significantly with existing merged cells or tables
            # (This is complex to check perfectly, provide a basic check)
            try:
                if table_range.merge_cells:
                    logger.warning(f"Target range {table_address} contains merged cells. Table insertion might fail or behave unexpectedly.")

                # Check for overlapping tables (ListObjects) - requires iterating existing tables
                # Wrap ListObjects access in its own try-except
                try:
                    if hasattr(sheet.api, 'ListObjects'): # Check if property exists first
                         for lo in sheet.api.ListObjects:
                             try: # Protect against errors querying individual table properties
                                 lo_range = sheet.range(lo.Range.Address)
                                 # Basic intersection check (can be refined)
                                 # if table_range.api.Column <= lo_range.last_cell.api.Column and ... (complex check)
                                 logger.debug(f"Found existing table '{lo.Name}' at {lo.Range.Address}. Overlap check not fully implemented.")
                             except Exception as single_lo_err:
                                logger.debug(f"Could not query details of an existing ListObject during pre-check: {single_lo_err}")
                             pass # Skipping complex overlap check for now
                    else:
                        logger.debug("ListObjects property not found on sheet.api, skipping table overlap check.")
                except Exception as list_objects_access_err:
                     logger.warning(f"Could not access sheet.api.ListObjects during pre-check (likely unsupported): {list_objects_access_err}")

            except Exception as overlap_check_err:
                # Catch other potential errors during pre-check (e.g., merge_cells check)
                logger.warning(f"Could not perform pre-check for overlaps: {overlap_check_err}")

            # --- Write header and data ---
            logger.debug("Writing table data to range...")
            # Use options(transpose=False) explicitly if needed, though default
            # Ensure data is list-of-lists: [columns] + normalized_rows
            full_data = [columns] + normalized_rows
            table_range.options(expand='table').value = full_data
            logger.debug("Data written successfully.")

            # --- Create the ListObject ---
            logger.debug(f"Attempting to add ListObject to range {table_address}...")
            lo = None # Initialize lo to None
            try:
                # Constants for ListObjects.Add: SourceType=xlSrcRange(1), Range, LinkSource, XlListObjectHasHeaders=xlYes(1)
                xlYes = 1 # xw.constants.XlYesNoGuess.xlYes
                xlSrcRange = 1 # xw.constants.ListObjectSourceType.xlSrcRange
                lo = sheet.api.ListObjects.Add(SourceType=xlSrcRange, Source=table_range.api, XlListObjectHasHeaders=xlYes)
                logger.debug("ListObject added successfully.")

                # --- Set Table Name and Style (Only if ListObject creation succeeded) ---
                if table_name:
                    try:
                        # Check if name already exists (more robust check needed if lo could be None earlier)
                        name_exists = False
                        for existing_lo in sheet.api.ListObjects:
                             # Check name and avoid comparing against itself if lo is valid
                            if existing_lo.Name == table_name and (lo is None or existing_lo.Range.Address != lo.Range.Address):
                                name_exists = True
                                break
                        if name_exists:
                           logger.warning(f"Table name '{table_name}' already exists on sheet '{sheet_name}'. Using default name.")
                           # Optionally generate a unique name here instead of relying on default
                        else:
                            logger.debug(f"Setting table name to '{table_name}'.")
                            lo.Name = table_name
                    except Exception as name_err:
                        logger.warning(f"Failed to set table name to '{table_name}': {name_err}. Using default name.")

                if table_style:
                    try:
                        logger.debug(f"Setting table style to '{table_style}'.")
                        lo.TableStyle = table_style
                    except Exception as style_err:
                        # Make warning more specific if style application failed vs table name
                        logger.warning(f"Failed to set table style to '{table_style}': {style_err}.")

            except Exception as table_api_err:
                # If ListObject creation failed, log clearly but don't raise an error here
                # The data has already been written.
                logger.warning(f"Could not create formal Excel Table (ListObject) for range {table_address} (Unsupported on this Excel version/OS or other error). Data written as plain range. Error: {table_api_err}")
                # lo remains None, so subsequent name/style setting is skipped implicitly

            # (Optional) Attempt an autofit on the written range (even if not a formal table)
            # Safe to skip if not supported
            try:
                table_range.columns.autofit()
            except Exception as autofit_err:
                logger.debug(f"Could not autofit table columns: {autofit_err}")

        except Exception as e:
            logger.error(f"Failed to insert table '{table_name}' into {sheet_name}!{start_cell}: {e}")
            # Fallback: Just format as a plain range
            try:
                logger.warning("Table creation failed or not supported. Applying minimal formatting fallback.")
                header_row = header_cell.resize(row_size=1, column_size=total_cols)
                # Make header bold, but safe-check if it's supported
                try:
                    header_row.font.bold = True
                except Exception as bold_err:
                    logger.debug(f"Could not set bold on header row: {bold_err}")
                # Attempt a fill color
                try:
                    header_row.color = _hex_argb_to_bgr_int("FFD9E1F2")
                except Exception as color_err:
                    logger.debug(f"Could not set header fill color: {color_err}")
                # Possibly skip borders entirely, as they're not cross-platform
            except Exception as fallback_err:
                logger.error(f"Fallback formatting also failed: {fallback_err}")
            raise RuntimeError(f"Failed to insert table '{table_name}': {e}") from e

    def append_table_rows(self, sheet_name: str, table_name: str, rows: List[List[Any]]) -> None:
        """
        Appends rows to an existing Excel table using COM ListRows.Add for incremental inserts.
        Ensures row data width matches table width.
        Falls back to plain range append if ListObjects are not supported in the environment.
        """
        sheet = self._require_sheet(sheet_name) # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)

        if not rows:
            logger.debug(f"No rows provided to append to table '{table_name}'. Skipping.")
            return

        logger.info(f"Appending {len(rows)} rows to table '{table_name}' on sheet '{sheet_name}'.")
        
        # Check if this environment supports ListObjects
        supports_listobjects = True
        try:
            # Try to access ListObjects property - this will fail on platforms without COM support
            # or older Excel versions
            hasattr(sheet.api, 'ListObjects')
        except Exception as env_check_err:
            supports_listobjects = False
            logger.warning(f"This environment doesn't appear to support Excel tables (ListObjects): {env_check_err}. Using plain range append fallback.")
        
        # If ListObjects not supported, use fallback approach immediately
        if not supports_listobjects:
            return self._append_table_rows_fallback(sheet_name, table_name, rows)
        
        try:
            # Locate the table (ListObject) by name using the sheet's API
            tbl = None
            try:
                # Accessing ListObjects by name might raise if not found
                tbl = sheet.api.ListObjects(table_name)
                logger.debug(f"Found table '{table_name}' via direct access.")
            except Exception as direct_err:
                # Check if this is a "property not found" error indicating no ListObjects support
                if "property" in str(direct_err).lower() and "not found" in str(direct_err).lower():
                    logger.warning(f"ListObjects property not found: {direct_err}. Using plain range append fallback.")
                    return self._append_table_rows_fallback(sheet_name, table_name, rows)
                    
                # Otherwise try alternate access method
                logger.debug(f"Direct access to table '{table_name}' failed, iterating...")
                try:
                    for lo in sheet.api.ListObjects:
                        if lo.Name == table_name:
                            tbl = lo
                            logger.debug(f"Found table '{table_name}' via iteration.")
                            break
                except Exception as iter_err:
                    logger.warning(f"Cannot iterate ListObjects: {iter_err}. Using plain range append fallback.")
                    return self._append_table_rows_fallback(sheet_name, table_name, rows)

            if tbl is None:
                # Try to find a named range with this table name before giving up
                try:
                    for name in self.book.names:
                        if name.name == table_name:
                            logger.info(f"Found '{table_name}' as a named range instead of table. Using plain range append.")
                            return self._append_table_rows_fallback(sheet_name, table_name, rows, named_range=name)
                except Exception as name_err:
                    logger.debug(f"Error checking for named range: {name_err}")
                    
                # If we reach here, we couldn't find the table or named range
                logger.error(f"Table '{table_name}' not found on sheet '{sheet_name}'")
                raise KeyError(f"Table '{table_name}' not found on sheet '{sheet_name}'")

            # Get table dimensions
            try:
                num_table_cols = tbl.ListColumns.Count
                table_range_addr = tbl.Range.Address
                logger.debug(f"Target table '{table_name}' found at {table_range_addr} with {num_table_cols} columns.")
            except Exception as table_dim_err:
                logger.warning(f"Could not get table dimensions: {table_dim_err}. Using plain range append fallback.")
                return self._append_table_rows_fallback(sheet_name, table_name, rows, table_obj=tbl)

            # --- Append rows one by one using ListRows.Add ---
            # This is generally safer for tables with formulas or special formatting
            added_count = 0
            for i, row_vals in enumerate(rows):
                try:
                    # Normalize the row data to match table width
                    # Note: _normalise_rows is now imported from core_defs
                    normalized_row = _normalise_rows([None]*num_table_cols, [row_vals])[0] # Use dummy header

                    # Add a new blank row to the table first
                    # The 'AlwaysInsert' parameter (if available, value=True) might help if adding to filtered table
                    try:
                        new_listrow = tbl.ListRows.Add(AlwaysInsert=True)
                    except TypeError: # Handle if AlwaysInsert arg is not supported
                        new_listrow = tbl.ListRows.Add()
                    except Exception as add_err:
                        # If adding row fails with COM error, fall back to direct append for this and remaining rows
                        logger.warning(f"ListRows.Add() failed: {add_err}. Switching to plain range append for remaining rows.")
                        # Get current row count and append the rest of the rows manually
                        remaining_rows = rows[i:]
                        return self._append_table_rows_fallback(sheet_name, table_name, remaining_rows, 
                                                              table_obj=tbl, already_added=added_count)

                    # Get the range corresponding to the new row
                    new_row_range = sheet.range(new_listrow.Range.Address.replace('$', ''))

                    # Write the normalized values into the new row's range
                    # Use .value to avoid interpreting data as formulas etc.
                    new_row_range.value = normalized_row
                    added_count += 1

                except Exception as row_add_err:
                    logger.error(f"Failed to append row {i+1} to table '{table_name}': {row_add_err}")
                    # Continue with next row
                    continue

            logger.info(f"Successfully appended {added_count}/{len(rows)} rows to table '{table_name}'.")
            return  # Success case

        except KeyError as ke:
            raise ke # Re-raise KeyErrors (table not found)
        except Exception as e:
            logger.error(f"Failed to append rows to table '{table_name}': {e}")
            # Try fallback before giving up completely
            try:
                logger.warning(f"Attempting fallback method for appending rows to '{table_name}'")
                return self._append_table_rows_fallback(sheet_name, table_name, rows)
            except Exception as fallback_err:
                logger.error(f"Fallback method also failed: {fallback_err}")
                raise RuntimeError(f"Failed to append rows to table '{table_name}': {e}") from e
                
    def _append_table_rows_fallback(self, sheet_name: str, table_name: str, rows: List[List[Any]], 
                                   table_obj=None, named_range=None, already_added: int = 0) -> None:
        """
        Fallback implementation to append rows when ListObjects API is not available.
        Works with plain ranges instead of formal Excel tables.
        
        Args:
            sheet_name: Sheet containing the table/range
            table_name: Name of the table/range
            rows: List of rows to append
            table_obj: Optional table object if already found
            named_range: Optional named range if found
            already_added: Number of rows already added (for reporting)
        """
        sheet = self._require_sheet(sheet_name)  # Ensures connection and sheet exists
        logger = logging.getLogger(__name__)
        
        if not rows:
            logger.debug("No rows to append in fallback method")
            return
            
        # 1. Identify the target range where the "table" data is
        target_range_addr = None
        
        # Case 1: We have a table object
        if table_obj is not None:
            try:
                target_range_addr = table_obj.Range.Address
                logger.debug(f"Using table object range: {target_range_addr}")
            except Exception as tbl_addr_err:
                logger.warning(f"Could not get table range address: {tbl_addr_err}")
                
        # Case 2: We have a named range
        if target_range_addr is None and named_range is not None:
            try:
                target_range_addr = named_range.refers_to_range.address
                logger.debug(f"Using named range: {target_range_addr}")
            except Exception as name_addr_err:
                logger.warning(f"Could not get named range address: {name_addr_err}")
                
        # Case 3: Search for the range by table name in existing ranges
        if target_range_addr is None:
            # Try various methods to find the range
            try:
                # Check named ranges
                for name in self.book.names:
                    if name.name == table_name:
                        try:
                            target_range_addr = name.refers_to_range.address
                            logger.debug(f"Found table as named range: {target_range_addr}")
                            break
                        except Exception:
                            pass
            except Exception as names_err:
                logger.debug(f"Error searching named ranges: {names_err}")
                
            # Check for a range with header text matching table_name
            if target_range_addr is None:
                logger.warning(f"Table '{table_name}' not found as a formal table or named range. Searching for header text...")
                # This is a more aggressive fallback - search for table-like data with headers
                try:
                    used_range = sheet.used_range
                    # Look for table headers in row 1
                    header_row = sheet.range('1:1').value
                    if isinstance(header_row, list):
                        # Check if any header cell contains the table name
                        for idx, cell_val in enumerate(header_row):
                            if cell_val and table_name.lower() in str(cell_val).lower():
                                # Found a likely header - determine table range
                                col_letter = get_column_letter(idx + 1)
                                last_row = used_range.last_cell.row
                                target_range_addr = f"{col_letter}1:{col_letter}{last_row}"
                                logger.debug(f"Found potential table by header text at column {col_letter}")
                                break
                except Exception as header_search_err:
                    logger.debug(f"Error during header search: {header_search_err}")
        
        # If we still don't have a range, we can't proceed
        if target_range_addr is None:
            raise KeyError(f"Could not locate table or range '{table_name}' for append operation")
            
        # 2. Determine the number of columns and last row of the target range
        try:
            # Get the range object
            target_range = sheet.range(target_range_addr)
            
            # Determine column count from target range
            if hasattr(target_range, 'columns') and hasattr(target_range.columns, 'count'):
                num_cols = target_range.columns.count
            else:
                # Alternative: calculate from address
                if ':' in target_range_addr:
                    try:
                        # Attempt to parse A1:B10 format
                        start_addr, end_addr = target_range_addr.split(':')
                        start_col = coordinate_from_string(start_addr.replace('$', ''))[0]
                        end_col = coordinate_from_string(end_addr.replace('$', ''))[0]
                        start_col_idx = column_index_from_string(start_col)
                        end_col_idx = column_index_from_string(end_col)
                        num_cols = end_col_idx - start_col_idx + 1
                    except Exception:
                        # Guess based on first row values
                        first_row_vals = target_range.value
                        if isinstance(first_row_vals, list) and len(first_row_vals) > 0:
                            if isinstance(first_row_vals[0], list):
                                num_cols = len(first_row_vals[0])
                            else:
                                num_cols = len(first_row_vals)
                        else:
                            num_cols = 1  # Default if all else fails
                else:
                    num_cols = 1  # Single cell range
            
            # Find the last row of the table
            if hasattr(target_range, 'rows') and hasattr(target_range.rows, 'count'):
                current_row_count = target_range.rows.count
            else:
                # Parse from address
                if ':' in target_range_addr:
                    try:
                        # Parse A1:B10 format
                        _, end_addr = target_range_addr.split(':')
                        end_row = coordinate_from_string(end_addr.replace('$', ''))[1]
                        current_row_count = end_row
                    except Exception:
                        # Fallback - get values and count rows
                        vals = target_range.value
                        if isinstance(vals, list):
                            if isinstance(vals[0], list):
                                current_row_count = len(vals)
                            else:
                                current_row_count = 1
                        else:
                            current_row_count = 1
                else:
                    current_row_count = 1  # Single cell
            
            # Get the starting range coordinates
            if ':' in target_range_addr:
                start_addr = target_range_addr.split(':')[0].replace('$', '')
            else:
                start_addr = target_range_addr.replace('$', '')
            
            start_col_letter = coordinate_from_string(start_addr)[0]
            start_col_idx = column_index_from_string(start_col_letter)
            
            logger.debug(f"Target range '{table_name}' has {num_cols} columns and {current_row_count} rows")
            
            # 3. Normalize and append each row
            successful_appends = 0
            for i, row_vals in enumerate(rows):
                try:
                    # Normalize the row data to match table width
                    normalized_row = _normalise_rows([None]*num_cols, [row_vals])[0]  # Use dummy header
                    
                    # Calculate the destination cell
                    next_row_num = current_row_count + 1 + i  # +1 because we append after the last row
                    append_cell = f"{start_col_letter}{next_row_num}"
                    
                    # Create a range for the new row
                    row_range = sheet.range(append_cell).resize(1, num_cols)
                    
                    # Write the values
                    row_range.value = normalized_row
                    successful_appends += 1
                    
                except Exception as append_err:
                    logger.error(f"Failed to append row {i+1} in fallback method: {append_err}")
                    continue
            
            logger.info(f"Fallback method: Successfully appended {successful_appends}/{len(rows)} rows " +
                      f"({already_added + successful_appends} total) to '{table_name}' using plain range append")
            
        except Exception as e:
            logger.error(f"Fallback append operation failed: {e}")
            raise RuntimeError(f"Could not append rows to '{table_name}' using either method: {e}") from e


# --------------------------------------------------------------------------
# Helper functions (moved outside class)
# --------------------------------------------------------------------------

# Color helpers (_hex_argb_to_bgr_int, _to_bgr, _bgr_int_to_argb_hex) moved to core_defs.py

def _safe_cell_style(cell) -> dict: # noqa: ANN001 – add types later if needed
    """Robustly extract font.bold + fill.color from any `xlwings.Range`."""
    # Return dict can be passed directly to ``openpyxl.cell.Cell.style``.
    style = {}
    logger = logging.getLogger(__name__) # Add logger for debug inside helper

    # Bold --------------------------------------------------------------------
    bold = None
    for probe in (
        lambda c: c.api.Font.Bold,   # fast on Windows
        lambda c: c.font.bold,       # works everywhere
    ):
        try:
            bold_val = probe(cell) # Use temp variable
            if bold_val is not None: # Check if probe returned a value
                bold = bold_val
                break # Stop probing if successful
        except Exception:
            pass # Ignore errors from individual probes, try next
    # Check final bold value *after* the loop
    if bold is not None:
        style.setdefault("font", {})["bold"] = bool(bold)

    # Fill --------------------------------------------------------------------
    color = None # This will store the final color value (likely BGR int)
    for probe in (
        lambda c: c.api.Interior.Color, # Usually BGR integer
        lambda c: c.color,              # Usually RGB tuple or None
    ):
        try:
            color_val = probe(cell)
            # Handle different return types from probes
            # Check for valid integer color (excluding None, 0=Black, 16777215=White)
            # We might want to capture black/white, adjust condition if needed
            # Let's assume 0 and 16777215 (Excel's default white index?) might be meaningful
            if isinstance(color_val, (int, float)) and color_val is not None:
                # Let's take any integer found
                color = int(color_val) # Assume BGR int if numeric
                break
            elif isinstance(color_val, tuple) and color_val is not None: # Check for RGB tuple
                # Convert RGB tuple to BGR int for consistency before converting to hex
                # Assuming color_val is (R, G, B)
                if len(color_val) == 3:
                    try:
                        r, g, b = map(int, color_val)
                        # Check if it's default black (0,0,0) or white (255,255,255) - maybe treat differently?
                        color = b << 16 | g << 8 | r # Convert to BGR integer
                        break
                    except (ValueError, TypeError):
                        logger.debug(f"Could not convert RGB tuple {color_val} to int.")
                        pass # Ignore tuple if conversion fails

            # else: ignore other types like None from c.color probe if Interior.Color already succeeded or failed
        except Exception as probe_err:
             logger.debug(f"Probe failed for cell {getattr(cell, 'address', 'unknown')}: {probe_err}")
             pass # Ignore errors from individual probes, try next

    # Check color value *after* the loop
    # Note: _bgr_int_to_argb_hex is now imported from core_defs
    if color is not None: # Check if color has *any* meaningful value found
        try:
            # Ensure conversion is safe. _bgr_int_to_argb_hex expects an int.
            hex_color = _bgr_int_to_argb_hex(color) # Convert BGR int to ARGB hex
            # Avoid adding default white 'FFFFFFFF' unless necessary? Depends on desired behavior.
            # if hex_color != 'FFFFFFFF': # Optional: filter out default white
            style.setdefault("fill", {})["start_color"] = hex_color
        except Exception as conv_err:
             # Log error during conversion if necessary, but don't add style if conversion fails
             logger.debug(f"Could not convert color value {color} for cell {getattr(cell, 'address', 'unknown')}: {conv_err}")

    return style


def _full_sheet_range(sheet: xw.Sheet) -> str:
    """Return A1-style full-used-range of a sheet."""
    logger = logging.getLogger(__name__) # Use logger inside helper
    try:
        # Use used_range which is generally reliable
        used_rng = sheet.used_range
        # Handle completely empty sheet case
        # Check value of top-left cell of used_range to distinguish empty from single-cell content
        top_left_val = used_rng.cells[0].value
        # Check if the used range is just A1 AND its value is None
        if used_rng.address == '$A$1' and top_left_val is None:
             # Return "A1" for a truly empty sheet (used range is A1, value is None)
             return "A1"
        else:
            # Otherwise, return the address of the used range (could be A1 if only A1 has content)
            return used_rng.address.replace('$', '')
    except Exception as e:
        logger.warning(f"Could not determine used range for sheet '{getattr(sheet, 'name', 'unknown')}', defaulting to A1: {e}")
        return "A1" # Default to A1 on error

# --- End of excel_ops.py ---